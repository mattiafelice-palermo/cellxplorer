"""Read structured Neware Excel exports into CellXplorer's raw model.

This module owns the Excel-specific source boundary for Spec 039: workbook metadata,
the programmed plan, the large ``record`` time series, and the small independent
cycle-summary validation.  Dispatch and normalized compatibility fields remain in
:mod:`parsing`; cache publication remains in :mod:`cache`.
"""
from __future__ import annotations

import math
import re
import zipfile
from contextlib import contextmanager
from datetime import time as datetime_time
from datetime import timedelta
from decimal import Decimal, InvalidOperation
from itertools import chain, islice
from numbers import Number
from pathlib import Path
from typing import Any, Iterator
from xml.etree import ElementTree

import numpy as np
import pandas as pd

try:
    import fastexcel as _fastexcel
except (ImportError, OSError):  # pragma: no cover - minimal source installs
    _fastexcel = None  # type: ignore[assignment]
if _fastexcel is not None and not callable(getattr(_fastexcel, "read_excel", None)):
    # A partially visible namespace package can occur in constrained source
    # environments.  Treat it as absent so the reference parser remains the
    # safe fallback instead of surfacing an import-shape AttributeError.
    _fastexcel = None  # type: ignore[assignment]
try:
    import python_calamine as _python_calamine
except (ImportError, OSError):  # pragma: no cover - minimal source installs
    _python_calamine = None  # type: ignore[assignment]
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException


EXCEL_PARSER_REVISION = 6


class _FastExcelFallback(Exception):
    """The optional columnar reader cannot represent this workbook safely."""


class _CalamineFallback(Exception):
    """The optional pandas/calamine reader cannot represent this workbook safely."""


_FAST_EXCEL_ERRORS: tuple[type[BaseException], ...] = (
    ImportError,
    OSError,
)
if _fastexcel is not None:
    _fast_excel_error = getattr(_fastexcel, "FastExcelError", None)
    if isinstance(_fast_excel_error, type):
        _FAST_EXCEL_ERRORS += (_fast_excel_error,)
_CALAMINE_ERRORS: tuple[type[BaseException], ...] = (
    ImportError,
    OSError,
    ValueError,
    RuntimeError,
    zipfile.BadZipFile,
)
if _python_calamine is not None:
    _calamine_error = getattr(_python_calamine, "CalamineError", None)
    if isinstance(_calamine_error, type):
        _CALAMINE_ERRORS += (_calamine_error,)

_TEST_METADATA_LABELS = {
    "start step id": "start_step_id",
    "volt. upper": "protection_voltage_upper_v",
    "volt. lower": "protection_voltage_lower_v",
    "builder": "builder",
    "remarks": "remarks",
    "start time": "start_time",
    "barcode": "barcode",
    "active material": "active_mass_mg",
    "nominal capacity": "nominal_capacity_mah",
    "record settings": "record_settings",
    "p/n": "part_number",
    "cycle count": "cycle_count",
    "voltage range": "voltage_range",
    "current range": "current_range",
}
_VALUE_GROUP_VALUE_OFFSET = 2

_PLAN_REQUIRED_HEADERS = ("Step Index", "Step Name")
_PLAN_HEADERS = (
    "Step Index",
    "Step Name",
    "Step Time(min)",
    "Voltage(V)",
    "C-rate(C)",
    "Current(mA)",
    "Cut-off voltage (V)",
    "Cut-off C-rate(C)",
    "Cut-off curr.(mA)",
    "Energy(Wh)",
    "-ΔV(V)",
    "Power(W)",
    "Resistance(mΩ)",
    "Capacity(mAh)",
    "Record settings",
    "Aux.CH recording condition",
    "Max Vi(V)",
    "Min Vi(V)",
    "Max Ti(℃)",
    "Min Ti(℃)",
    "Segment record1",
    "Segment record2",
    "Current range (mA)",
)

_CYCLE_REQUIRED_HEADERS = (
    "Cycle Index",
    "Chg. Cap.(mAh)",
    "DChg. Cap.(mAh)",
    "Chg. Time(min)",
    "DChg. Time(min)",
)
_CYCLE_OPTIONAL_HEADERS = (
    "Chg.-DChg. Eff(%)",
    "Chg. Energy(Wh)",
    "DChg. Energy(Wh)",
)
_CYCLE_HEADERS = _CYCLE_REQUIRED_HEADERS + _CYCLE_OPTIONAL_HEADERS

# Neware changes labels and displayed units between export surfaces and
# software versions.  These aliases are deliberately scoped to the worksheet
# that owns the field: ``record`` may use Cycle ID while ``step`` uses Cycle
# Index in the same workbook.  A resolver rejects a workbook that contains
# both aliases for one semantic field instead of silently picking one.
_HEADER_ALIASES: dict[str, dict[str, tuple[str, ...]]] = {
    "record": {
        "Cycle Index": ("Cycle Index", "Cycle ID"),
        "Step Index": ("Step Index", "Step ID"),
        "Time(min)": ("Time(min)", "Time"),
        "Total Time(min)": ("Total Time(min)", "Total Time"),
        "Power(W)": ("Power(W)", "Power(kW)"),
    },
    "step": {
        "Cycle Index": ("Cycle Index", "Cycle ID"),
        "Step Index": ("Step Index", "Step ID"),
        "Step Time(min)": ("Step Time(min)", "Step Time"),
        "Energy(Wh)": ("Energy(Wh)", "Energy(kWh)"),
    },
    "cycle": {
        "Cycle Index": ("Cycle Index", "Cycle ID"),
        "Chg. Time(min)": ("Chg. Time(min)", "Chg. Time"),
        "DChg. Time(min)": ("DChg. Time(min)", "DChg. Time"),
        "Chg. Energy(Wh)": ("Chg. Energy(Wh)", "Chg. Energy(kWh)"),
        "DChg. Energy(Wh)": ("DChg. Energy(Wh)", "DChg. Energy(kWh)"),
    },
}

_PLAN_HEADER_ALIASES: dict[str, tuple[str, ...]] = {
    "Step Time(min)": ("Step Time(min)", "Step Time(hh:mm:ss.ms)", "Step Time"),
    "Energy(Wh)": ("Energy(Wh)", "Energy(kWh)"),
    "Power(W)": ("Power(W)", "Power(kW)"),
}

class NewareExcelError(ValueError):
    """Base class for bounded Neware Excel parser errors."""


class UnsupportedNewareExcelError(NewareExcelError):
    """The workbook is not the supported structured Neware export."""


class InvalidNewareExcelError(NewareExcelError):
    """The workbook resembles the supported export but is unsafe to map."""


REQUIRED_RECORD_HEADERS = (
    "DataPoint",
    "Cycle Index",
    "Step Index",
    "Step Type",
    "Time(min)",
    "Total Time(min)",
    "Current(mA)",
    "Voltage(V)",
    "Chg. Cap.(mAh)",
    "DChg. Cap.(mAh)",
    "Date",
    "Power(W)",
)

OPTIONAL_RECORD_HEADERS = {
    "Capacity(mAh)": "capacity_mah",
    "Spec. Cap.(mAh/g)": "specific_capacity_mah_g",
    "Chg. Spec. Cap.(mAh/g)": "charge_specific_capacity_mah_g",
    "DChg. Spec. Cap.(mAh/g)": "discharge_specific_capacity_mah_g",
}

STEP_HEADERS = (
    "Cycle Index",
    "Step Index",
    "Step Number",
    "Step Type",
    "Step Time(min)",
    "Oneset Date",
    "End Date",
    "Capacity(mAh)",
    "Energy(Wh)",
    "Oneset Volt.(V)",
    "End Voltage(V)",
)

_RECORD_OUTPUT_COLUMNS = {
    "DataPoint": "record_index",
    "Cycle Index": "cycle",
    "Step Index": "step_index",
    "Time(min)": "time_s",
    "Total Time(min)": "total_time_s",
    "Current(mA)": "current_ma",
    "Voltage(V)": "voltage_v",
    "Chg. Cap.(mAh)": "charge_capacity_mah",
    "DChg. Cap.(mAh)": "discharge_capacity_mah",
    "Date": "timestamp",
    "Power(W)": "power_w",
}

_STATUS_ALIASES = {
    "rest": "Rest",
    "cc chg": "CC_Chg",
    "cc dchg": "CC_DChg",
    "cv chg": "CV_Chg",
    "cccv chg": "CCCV_Chg",
    "cccv dchg": "CCCV_DChg",
}

_INT_COLUMNS = ("record_index", "cycle", "step", "step_index")
_FLOAT_COLUMNS = (
    "time_s",
    "total_time_s",
    "voltage_v",
    "current_ma",
    "charge_capacity_mah",
    "discharge_capacity_mah",
    "charge_energy_mwh",
    "discharge_energy_mwh",
    "power_w",
    "capacity_mah",
    "specific_capacity_mah_g",
    "charge_specific_capacity_mah_g",
    "discharge_specific_capacity_mah_g",
)


def _normalize_text(value: object) -> str:
    """Normalize text for deterministic lookup, without fuzzy matching."""

    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).strip()).casefold()


def _normalize_label(value: object) -> str:
    return re.sub(r"[:：]\s*$", "", _normalize_text(value))


def _is_blank(value: object) -> bool:
    if value is None:
        return True
    return _normalize_text(value) in {"", "-", "–", "—", "n/a", "na"}


def _value_text(value: object) -> str | None:
    if _is_blank(value):
        return None
    if isinstance(value, pd.Timestamp):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    return str(value).strip()


def _format_number(value: float) -> str:
    if float(value).is_integer():
        return str(int(value))
    return format(float(value), ".15g")


_QUANTITY_RE = re.compile(
    r"^\s*([+-]?(?:\d+(?:\.\d*)?|\.\d+)(?:[eE][+-]?\d+)?)\s*([A-Za-zµμΩω℃]+)\s*$"
)
_PLAIN_NUMBER_RE = re.compile(
    r"^\s*[+-]?(?:\d+(?:\.\d*)?|\.\d+)(?:[eE][+-]?\d+)?\s*$"
)
_DURATION_RE = re.compile(
    r"^(?P<hours>\d+):(?P<minutes>[0-5]\d):(?P<seconds>[0-5]\d)"
    r"(?:\.(?P<fraction>\d{1,6}))?$"
)


def _clock_duration_seconds(value: object) -> float:
    """Parse Neware's unbounded ``H+:MM:SS[.ffffff]`` duration values.

    Neware sometimes writes elapsed durations as strings and sometimes lets
    openpyxl decode formatted Excel duration cells into ``timedelta`` or
    ``time`` objects.  The first component is intentionally unbounded: it is
    elapsed hours, not a clock hour that rolls over at 24.
    """

    if isinstance(value, timedelta):
        seconds = value.total_seconds()
        if math.isfinite(seconds) and seconds >= 0.0:
            return float(seconds)
        raise ValueError("negative or non-finite duration")
    if isinstance(value, datetime_time):
        return (
            value.hour * 3600.0
            + value.minute * 60.0
            + value.second
            + value.microsecond / 1_000_000.0
        )
    if isinstance(value, Number) and not isinstance(value, bool):
        raise ValueError("numeric values are ambiguous under a unitless duration header")

    text = str(value).strip() if value is not None else ""
    match = _DURATION_RE.fullmatch(text)
    if match is None:
        raise ValueError("duration must use H+:MM:SS")
    fraction = match.group("fraction") or ""
    fraction_seconds = int(fraction.ljust(6, "0")) / 1_000_000.0 if fraction else 0.0
    return (
        int(match.group("hours")) * 3600.0
        + int(match.group("minutes")) * 60.0
        + int(match.group("seconds"))
        + fraction_seconds
    )


def _unitless_or_minutes_seconds(
    value: object,
    *,
    source_header: str,
    canonical_header: str,
) -> float:
    """Convert a duration according to the exact resolved header alias."""

    if _normalize_text(source_header) == _normalize_text(canonical_header):
        try:
            number = float(value)
        except (TypeError, ValueError) as exc:
            raise ValueError("duration is not numeric minutes") from exc
        if not math.isfinite(number):
            raise ValueError("duration is not finite")
        return number * 60.0
    return _clock_duration_seconds(value)


def _is_clock_duration_source(source_header: str, canonical_header: str) -> bool:
    """Return whether a resolved duration header contains an elapsed clock."""

    return _normalize_text(source_header) != _normalize_text(canonical_header)


def _power_to_watts_factor(source_header: str) -> float:
    return 1000.0 if _normalize_text(source_header) == "power(kw)" else 1.0


def _energy_to_mwh_factor(source_header: str) -> float:
    return 1_000_000.0 if "(kwh)" in _normalize_text(source_header) else 1_000.0


def _unit_name(value: str) -> str:
    normalized = value.strip().casefold().replace("μ", "µ")
    return {
        "milligram": "mg",
        "milligrams": "mg",
        "milliamphour": "mah",
        "milliamphours": "mah",
        "millivolt": "mv",
        "millivolts": "mv",
        "millisecond": "ms",
        "milliseconds": "ms",
        "second": "s",
        "seconds": "s",
        "minute": "min",
        "minutes": "min",
    }.get(normalized, normalized)


def _quantity(
    value: object,
    *,
    label: str,
    expected_unit: str,
    required_unit: bool = False,
) -> float | None:
    """Parse a quantity using its declared unit, never a first-number guess."""

    if _is_blank(value):
        return None
    if isinstance(value, (int, float, np.integer, np.floating)) and not isinstance(value, bool):
        number = float(value)
        if math.isfinite(number):
            return number
        raise InvalidNewareExcelError(f"Neware Excel {label} has a non-finite value.")

    text = str(value).strip()
    match = _QUANTITY_RE.fullmatch(text)
    if match is None:
        if not required_unit and _PLAIN_NUMBER_RE.fullmatch(text):
            number = float(text)
            if math.isfinite(number):
                return number
        raise InvalidNewareExcelError(
            f"Neware Excel {label} must include a valid {expected_unit} quantity."
        )
    actual_unit = _unit_name(match.group(2))
    wanted_unit = _unit_name(expected_unit)
    if actual_unit != wanted_unit:
        raise InvalidNewareExcelError(
            f"Neware Excel {label} has contradictory unit {match.group(2)}; expected {expected_unit}."
        )
    number = float(match.group(1))
    if not math.isfinite(number):
        raise InvalidNewareExcelError(f"Neware Excel {label} has a non-finite value.")
    return number


def _plan_quantity(value: object, *, label: str, unit: str) -> float | None:
    return _quantity(value, label=label, expected_unit=unit, required_unit=False)


def _integer_value(value: object, *, label: str, required: bool = False) -> int | None:
    if _is_blank(value):
        if required:
            raise InvalidNewareExcelError(f"Neware Excel {label} is required.")
        return None
    number = _plan_quantity(value, label=label, unit="number")
    if number is None or not float(number).is_integer():
        raise InvalidNewareExcelError(f"Neware Excel {label} must be an integer.")
    return int(number)


def _metadata_quantity(value: object, *, label: str, unit: str) -> float | None:
    return _quantity(value, label=label, expected_unit=unit, required_unit=True)


def _metadata_timestamp(value: object, *, label: str) -> str | None:
    if _is_blank(value):
        return None
    try:
        timestamp = _coerce_timestamp(value)
    except (TypeError, ValueError, OverflowError) as exc:
        raise InvalidNewareExcelError(f"Neware Excel {label} has an invalid timestamp.") from exc
    return timestamp.strftime("%Y-%m-%d %H:%M:%S")


def _coerce_timestamp(value: object) -> pd.Timestamp:
    """Accept verified date-like values, but never bare Excel serial numbers."""

    if isinstance(value, Number) or (
        isinstance(value, str) and _PLAIN_NUMBER_RE.fullmatch(value.strip())
    ):
        raise ValueError("bare numeric timestamp")
    timestamp = pd.Timestamp(value)
    if pd.isna(timestamp) or timestamp.tz is not None:
        raise ValueError("invalid timestamp")
    return timestamp


def _record_settings(value: object, *, label: str = "Record settings") -> dict[str, float | str] | None:
    if _is_blank(value):
        return None
    text = str(value).strip()
    parts = text.split("/")
    if len(parts) != 3:
        raise InvalidNewareExcelError(
            f"Neware Excel {label} must use interval/voltage-delta/current-delta settings."
        )
    parsed: list[float] = []
    expected_units = ("s", "V", "mA")
    for part, expected in zip(parts, expected_units):
        try:
            parsed_value = _quantity(
                part,
                label=label,
                expected_unit=expected,
                required_unit=True,
            )
        except InvalidNewareExcelError:
            if expected != "s":
                raise
            # Neware's export UI commonly writes the record interval as
            # ``60000ms`` even though CellXplorer's protocol model stores
            # seconds.  This is an explicit unit alias, not a first-number
            # guess.
            parsed_value = _quantity(
                part,
                label=label,
                expected_unit="ms",
                required_unit=True,
            )
            if parsed_value is not None:
                parsed_value /= 1000.0
        if parsed_value is None:
            raise InvalidNewareExcelError(f"Neware Excel {label} contains an empty setting.")
        parsed.append(parsed_value)
    return {
        "raw": text,
        "interval_s": parsed[0],
        "voltage_delta_v": parsed[1],
        "current_delta_ma": parsed[2],
    }


def _rows(sheet: Any) -> list[tuple[object, ...]]:
    reset_dimensions = getattr(sheet, "reset_dimensions", None)
    if callable(reset_dimensions):
        reset_dimensions()
    return [tuple(row) for row in sheet.iter_rows(values_only=True)]


def _find_labeled_value(rows: list[tuple[object, ...]], label: str) -> object | None:
    """Return the value in a verified Neware label/value group.

    The exported information layouts place labels and values two columns apart
    (A/C, D/F, G/I; and the same positional groups on the ``unit`` sheet).
    Reading that fixed value slot keeps a blank value bounded to its own group;
    scanning arbitrary later cells can accidentally consume the next group's
    unsupported label.
    """

    wanted = _normalize_label(label)
    for row in rows:
        for index, value in enumerate(row):
            if _normalize_label(value) != wanted:
                continue
            value_index = index + _VALUE_GROUP_VALUE_OFFSET
            if value_index >= len(row):
                return None
            candidate = row[value_index]
            return None if _is_blank(candidate) else candidate
    return None


def _find_step_plan(
    rows: list[tuple[object, ...]],
) -> tuple[int, dict[str, tuple[int, str]]]:
    marker_row: int | None = None
    for row_number, row in enumerate(rows):
        if any(_normalize_label(value) == "step plan" for value in row):
            marker_row = row_number
            break
    if marker_row is None:
        raise InvalidNewareExcelError("Neware Excel test sheet has no Step plan marker.")

    for row_number in range(marker_row + 1, len(rows)):
        row = rows[row_number]
        headers: dict[str, tuple[int, str]] = {}
        for index, value in enumerate(row):
            normalized = _normalize_text(value)
            if normalized:
                original = str(value).strip()
                if normalized in headers:
                    raise InvalidNewareExcelError(
                        "Neware Excel test sheet has an ambiguous normalized Step plan header: "
                        f"{original}."
                    )
                headers[normalized] = (index, original)
        if all(_normalize_text(header) in headers for header in _PLAN_REQUIRED_HEADERS):
            return row_number, headers
    raise InvalidNewareExcelError("Neware Excel test sheet has no Step plan header row.")


def _original_key(header: str) -> str:
    words = re.findall(r"[A-Za-z0-9]+", header)
    return "".join(word[:1].upper() + word[1:] for word in words) or "Value"


def _plan_header_aliases(header: str) -> tuple[str, ...]:
    return _PLAN_HEADER_ALIASES.get(header, (header,))


def _plan_binding(
    headers: dict[str, tuple[int, str]],
    header: str,
) -> tuple[int, str] | None:
    matches = [
        headers[_normalize_text(alias)]
        for alias in _plan_header_aliases(header)
        if _normalize_text(alias) in headers
    ]
    if len(matches) > 1:
        raise InvalidNewareExcelError(
            "Neware Excel test sheet has ambiguous Step plan aliases for "
            f"{header}."
        )
    return matches[0] if matches else None


def _plan_value(
    row: tuple[object, ...],
    headers: dict[str, tuple[int, str]],
    header: str,
) -> object:
    binding = _plan_binding(headers, header)
    if binding is None:
        return None
    index, _source = binding
    return row[index] if index < len(row) else None


def _plan_source_header(headers: dict[str, tuple[int, str]], header: str) -> str:
    binding = _plan_binding(headers, header)
    return binding[1] if binding is not None else header


def _step_type_id(step_name: object) -> int:
    from .protocol import STEP_TYPES

    aliases = {
        "cc chg": "cc charge",
        "cc dchg": "cc discharge",
        "cv chg": "cv charge",
        "cccv chg": "cccv charge",
        "cccv dchg": "cccv discharge",
    }
    normalized = _normalize_text(step_name)
    canonical = aliases.get(normalized, normalized)
    matches = [
        type_id
        for type_id, (label, _direction) in STEP_TYPES.items()
        if _normalize_text(label) == canonical
    ]
    if len(matches) != 1:
        raise InvalidNewareExcelError(
            f"Neware Excel has an unsupported programmed Step Name: {step_name}."
        )
    return matches[0]


def _parse_loop_label(value: object, *, prefix: str, label: str) -> int:
    text = _value_text(value) or ""
    match = re.fullmatch(rf"{re.escape(prefix)}\s*:\s*(\d+)", text, flags=re.IGNORECASE)
    if match is None:
        raise InvalidNewareExcelError(
            f"Neware Excel {label} must use the labelled {prefix} form."
        )
    return int(match.group(1))


def _parse_test_information(
    rows: list[tuple[object, ...]],
) -> tuple[dict[str, object], int, dict[str, tuple[int, str]]]:
    marker_row, headers = _find_step_plan(rows)

    raw: dict[str, object] = {}
    for label, key in _TEST_METADATA_LABELS.items():
        value = _find_labeled_value(rows[:marker_row], label)
        raw[key] = value

    info: dict[str, object] = {
        "raw": raw,
        "start_step_id": _integer_value(raw["start_step_id"], label="Start step ID"),
        "protection_voltage_upper_v": _metadata_quantity(
            raw["protection_voltage_upper_v"], label="Volt. upper", unit="V"
        ),
        "protection_voltage_lower_v": _metadata_quantity(
            raw["protection_voltage_lower_v"], label="Volt. lower", unit="V"
        ),
        "builder": _value_text(raw["builder"]),
        "remarks": _value_text(raw["remarks"]),
        "start_time": _metadata_timestamp(raw["start_time"], label="Start time"),
        "barcode": _value_text(raw["barcode"]),
        "active_mass_mg": _metadata_quantity(
            raw["active_mass_mg"], label="Active material", unit="mg"
        ),
        "nominal_capacity_mah": _metadata_quantity(
            raw["nominal_capacity_mah"], label="Nominal capacity", unit="mAh"
        ),
        "record_settings": _record_settings(raw["record_settings"]),
        "part_number": _value_text(raw["part_number"]),
        "cycle_count": _value_text(raw["cycle_count"]),
        "voltage_range": _value_text(raw["voltage_range"]),
        "current_range": _value_text(raw["current_range"]),
        "marker_row": marker_row,
        "plan_headers": headers,
    }
    return info, marker_row, headers


def _declared_record_interval_seconds(workbook: Any) -> float | None:
    """Read the optional declared record cadence without making it required."""

    test_sheet = _sheet_by_name(workbook, "test", required=False)
    if test_sheet is None:
        return None
    try:
        info, _header_row, _headers = _parse_test_information(_rows(test_sheet))
    except NewareExcelError:
        return None
    settings = info.get("record_settings")
    if not isinstance(settings, dict):
        return None
    interval_s = settings.get("interval_s")
    return float(interval_s) if interval_s is not None else None


def _parse_unit_original(sheet: Any | None) -> tuple[dict[str, object], str | None]:
    if sheet is None:
        return {}, None
    rows = _rows(sheet)
    original: dict[str, object] = {}
    workbook_name = _value_text(rows[0][0]) if rows and rows[0] else None
    if workbook_name:
        original["WorkbookName"] = {"Value": workbook_name}

    for label, key in (("Start time", "StartTime"), ("End time", "EndTime")):
        value = _find_labeled_value(rows, label)
        if not _is_blank(value):
            original[key] = {"Value": _metadata_timestamp(value, label=label)}

    for row_index, row in enumerate(rows):
        if _normalize_label(row[0] if row else None) != "time":
            continue
        if row_index + 1 >= len(rows):
            break
        units = rows[row_index + 1]
        column_units: dict[str, str] = {}
        for index, header in enumerate(row):
            if _is_blank(header) or index >= len(units) or _is_blank(units[index]):
                continue
            column_units[str(header).strip()] = str(units[index]).strip()
        if column_units:
            original["ColumnUnits"] = {key: {"Value": value} for key, value in column_units.items()}
        break
    return original, (workbook_name if workbook_name else None)


def _parse_programmed_plan(
    rows: list[tuple[object, ...]],
    header_row: int,
    headers: dict[str, tuple[int, str]],
    info: dict[str, object],
) -> tuple[dict[str, dict[str, str]], dict[str, dict[str, object]]]:
    rows_by_index: dict[int, tuple[object, ...]] = {}
    for row in rows[header_row + 1 :]:
        if not any(not _is_blank(value) for value in row):
            continue
        step_index = _integer_value(
            _plan_value(row, headers, "Step Index"),
            label="Step Index",
            required=True,
        )
        if step_index is None or step_index <= 0:
            raise InvalidNewareExcelError("Neware Excel Step Index must be positive.")
        if step_index in rows_by_index:
            raise InvalidNewareExcelError(
                f"Neware Excel Step Index {step_index} is duplicated."
            )
        rows_by_index[step_index] = row

    expected = list(range(1, len(rows_by_index) + 1))
    if sorted(rows_by_index) != expected:
        raise InvalidNewareExcelError(
            "Neware Excel Step Index values must form a contiguous plan starting at 1."
        )

    protection_upper = info.get("protection_voltage_upper_v")
    protection_lower = info.get("protection_voltage_lower_v")
    global_settings = info.get("record_settings")
    steps: dict[str, dict[str, str]] = {}
    original: dict[str, dict[str, object]] = {}

    for step_index in expected:
        row = rows_by_index[step_index]
        step_name = _plan_value(row, headers, "Step Name")
        type_id = _step_type_id(step_name)
        step: dict[str, str] = {"Step_Type": str(type_id)}

        source_values: dict[str, object] = {}
        for header in _PLAN_HEADERS:
            value = _plan_value(row, headers, header)
            if not _is_blank(value):
                source_header = _plan_source_header(headers, header)
                source_values[_original_key(source_header)] = {"Value": _value_text(value)}

        if type_id == 5:
            step["Limit.Other.Start_Step.Value"] = str(
                _parse_loop_label(
                    _plan_value(row, headers, "Step Time(min)"),
                    prefix="Start step ID",
                    label=f"Step {step_index} Step Time(min)",
                )
            )
            step["Limit.Other.Cycle_Count.Value"] = str(
                _parse_loop_label(
                    _plan_value(row, headers, "Voltage(V)"),
                    prefix="Cycle count",
                    label=f"Step {step_index} Voltage(V)",
                )
            )
        elif type_id != 6:
            current = _plan_quantity(
                _plan_value(row, headers, "Current(mA)"),
                label=f"Step {step_index} Current(mA)",
                unit="mA",
            )
            rate = _plan_quantity(
                _plan_value(row, headers, "C-rate(C)"),
                label=f"Step {step_index} C-rate(C)",
                unit="C",
            )
            target_voltage = _plan_quantity(
                _plan_value(row, headers, "Voltage(V)"),
                label=f"Step {step_index} Voltage(V)",
                unit="V",
            )
            stop_voltage = _plan_quantity(
                _plan_value(row, headers, "Cut-off voltage (V)"),
                label=f"Step {step_index} Cut-off voltage (V)",
                unit="V",
            )
            stop_current = _plan_quantity(
                _plan_value(row, headers, "Cut-off curr.(mA)"),
                label=f"Step {step_index} Cut-off curr.(mA)",
                unit="mA",
            )
            stop_rate = _plan_quantity(
                _plan_value(row, headers, "Cut-off C-rate(C)"),
                label=f"Step {step_index} Cut-off C-rate(C)",
                unit="C",
            )
            nominal_capacity = info.get("nominal_capacity_mah")
            if stop_current is None and stop_rate is not None and nominal_capacity is not None:
                stop_current = abs(stop_rate * float(nominal_capacity))
            step_time_value = _plan_value(row, headers, "Step Time(min)")
            step_time_source = _plan_source_header(headers, "Step Time(min)")
            if _normalize_text(step_time_source) == _normalize_text("Step Time(min)"):
                step_time_min = _plan_quantity(
                    step_time_value,
                    label=f"Step {step_index} Step Time(min)",
                    unit="min",
                )
            elif _is_blank(step_time_value):
                step_time_min = None
            else:
                try:
                    step_time_min = _clock_duration_seconds(step_time_value) / 60.0
                except (TypeError, ValueError, OverflowError) as exc:
                    raise InvalidNewareExcelError(
                        f"Neware Excel Step {step_index} Step Time must use H+:MM:SS."
                    ) from exc
            if current is not None:
                step["Limit.Main.Curr.Value"] = _format_number(current)
            if rate is not None:
                step["Limit.Main.Rate.Value"] = _format_number(rate)
            if target_voltage is not None:
                step["Limit.Main.Volt.Value"] = _format_number(target_voltage * 10000.0)
            if stop_voltage is not None:
                step["Limit.Main.Stop_Volt.Value"] = _format_number(stop_voltage * 10000.0)
            if stop_current is not None:
                step["Limit.Main.Stop_Curr.Value"] = _format_number(stop_current)
            if step_time_min is not None:
                step["Limit.Main.Time.Value"] = _format_number(step_time_min * 60.0 * 1000.0)

            settings_value = _plan_value(row, headers, "Record settings")
            settings = _record_settings(settings_value) if not _is_blank(settings_value) else global_settings
            if settings is not None:
                step["Record.Main.Time.Value"] = _format_number(float(settings["interval_s"]) * 1000.0)
                step["Record.Main.Volt.Value"] = _format_number(float(settings["voltage_delta_v"]) * 10000.0)
                source_values["RecordCurrentDelta"] = {
                    "Value": _format_number(float(settings["current_delta_ma"]))
                }
            if protection_upper is not None:
                step["Protect.Main.Volt.Upper.Value"] = _format_number(float(protection_upper) * 10000.0)
            if protection_lower is not None:
                step["Protect.Main.Volt.Lower.Value"] = _format_number(float(protection_lower) * 10000.0)

        steps[f"Step{step_index}"] = step
        original[f"Step{step_index}"] = source_values

    return steps, original


def _path(path: str | Path) -> Path:
    return Path(path)


def _load(path: Path):
    try:
        return load_workbook(
            filename=path,
            read_only=True,
            data_only=True,
            keep_links=False,
        )
    except (InvalidFileException, OSError, ValueError, KeyError, zipfile.BadZipFile, ElementTree.ParseError) as exc:
        raise InvalidNewareExcelError(
            "Could not read the Neware Excel workbook."
        ) from exc
    except Exception as exc:  # openpyxl can surface parser-specific XML exceptions.
        raise InvalidNewareExcelError(
            "Could not read the Neware Excel workbook."
        ) from exc


@contextmanager
def _open(path: Path) -> Iterator[Any]:
    workbook = _load(path)
    try:
        yield workbook
    finally:
        workbook.close()


def _sheet_by_name(workbook: Any, name: str, *, required: bool) -> Any | None:
    wanted = _normalize_text(name)
    matches = [sheet for sheet in workbook.worksheets if _normalize_text(sheet.title) == wanted]
    if len(matches) > 1:
        raise InvalidNewareExcelError(
            f"Neware Excel workbook has ambiguous worksheet name: {name}."
        )
    if matches:
        return matches[0]
    if required:
        raise UnsupportedNewareExcelError(
            f"Not a recognized Neware Excel export: required {name} sheet is missing."
        )
    return None


def _header_map(sheet: Any) -> dict[str, tuple[int, str]]:
    # Some Neware exports declare ``<dimension ref=\"A1\"/>`` even though the
    # worksheet contains a full rectangular table.  Read-only openpyxl trusts
    # that declaration and would otherwise expose only column A.  Resetting
    # dimensions makes it scan the actual worksheet cells while preserving the
    # bounded read-only iteration path.
    reset_dimensions = getattr(sheet, "reset_dimensions", None)
    if callable(reset_dimensions):
        reset_dimensions()
    rows = sheet.iter_rows(min_row=1, max_row=1, values_only=True)
    try:
        header_row = next(rows)
    except StopIteration as exc:
        raise UnsupportedNewareExcelError(
            f"Neware Excel {sheet.title} sheet is empty."
        ) from exc

    result: dict[str, tuple[int, str]] = {}
    for index, value in enumerate(header_row):
        original = "" if value is None else str(value).strip()
        normalized = _normalize_text(value)
        if not normalized:
            continue
        if normalized in result:
            raise InvalidNewareExcelError(
                f"Neware Excel {sheet.title} sheet has an ambiguous normalized header: {original}."
            )
        result[normalized] = (index, original)
    return result


def _header_aliases(sheet_name: str, header: str) -> tuple[str, ...]:
    return _HEADER_ALIASES.get(sheet_name, {}).get(header, (header,))


def _resolve_header(
    headers: dict[str, tuple[int, str]],
    header: str,
    *,
    sheet_name: str,
) -> tuple[int, str] | None:
    matches = [
        headers[_normalize_text(alias)]
        for alias in _header_aliases(sheet_name, header)
        if _normalize_text(alias) in headers
    ]
    if len(matches) > 1:
        names = ", ".join(match[1] for match in matches)
        raise InvalidNewareExcelError(
            f"Neware Excel {sheet_name} sheet has ambiguous aliases for {header}: {names}."
        )
    return matches[0] if matches else None


def _require_columns(
    headers: dict[str, tuple[int, str]],
    required: tuple[str, ...],
    *,
    sheet_name: str,
) -> dict[str, tuple[int, str]]:
    resolved: dict[str, tuple[int, str]] = {}
    missing: list[str] = []
    for name in required:
        binding = _resolve_header(headers, name, sheet_name=sheet_name)
        if binding is None:
            missing.append(name)
        else:
            resolved[_normalize_text(name)] = binding
    if missing:
        if sheet_name == "record":
            raise UnsupportedNewareExcelError(
                "Neware Excel export is missing required record column: "
                f"{missing[0]}."
            )
        raise InvalidNewareExcelError(
            f"Neware Excel {sheet_name} sheet is missing required column: {missing[0]}."
        )
    return resolved


def _optional_columns(
    headers: dict[str, tuple[int, str]],
    optional: tuple[str, ...],
    *,
    sheet_name: str,
) -> dict[str, tuple[int, str]]:
    resolved: dict[str, tuple[int, str]] = {}
    for name in optional:
        binding = _resolve_header(headers, name, sheet_name=sheet_name)
        if binding is not None:
            resolved[_normalize_text(name)] = binding
    return resolved


def _record_number(row_number: int) -> int:
    # Worksheet row 2 is source record 1.  This fallback is used when DataPoint
    # itself is the invalid field and therefore cannot identify the record.
    return max(1, row_number - 1)


def _invalid_record(row_number: int, column: str) -> InvalidNewareExcelError:
    return InvalidNewareExcelError(
        f"Neware Excel record {_record_number(row_number)} has an invalid {column} value."
    )


def _number(value: object, *, row_number: int, column: str) -> float:
    if value is None or (isinstance(value, str) and not value.strip()):
        raise _invalid_record(row_number, column)
    try:
        number = float(value)
    except (TypeError, ValueError) as exc:
        raise _invalid_record(row_number, column) from exc
    if not math.isfinite(number):
        raise _invalid_record(row_number, column)
    return number


def _optional_number(value: object, *, row_number: int, column: str) -> float:
    if value is None or (isinstance(value, str) and not value.strip()):
        return math.nan
    return _number(value, row_number=row_number, column=column)


def _integer(value: object, *, row_number: int, column: str) -> int:
    number = _number(value, row_number=row_number, column=column)
    int_info = np.iinfo(np.int64)
    if (
        not number.is_integer()
        or number < int_info.min
        or number >= 2**63
    ):
        raise _invalid_record(row_number, column)
    return int(number)


def _timestamp(value: object, *, row_number: int, column: str) -> pd.Timestamp:
    if value is None or (isinstance(value, str) and not value.strip()):
        raise _invalid_record(row_number, column)
    try:
        timestamp = _coerce_timestamp(value)
    except (TypeError, ValueError, OverflowError) as exc:
        raise _invalid_record(row_number, column) from exc
    return timestamp


def _normalize_status(value: object, *, row_number: int, column: str = "Step Type") -> str:
    normalized = _normalize_text(value)
    if not normalized or normalized not in _STATUS_ALIASES:
        raise _invalid_record(row_number, column)
    return _STATUS_ALIASES[normalized]


def _parse_records(sheet: Any, headers: dict[str, tuple[int, str]]) -> list[dict[str, object]]:
    required = _require_columns(headers, REQUIRED_RECORD_HEADERS, sheet_name="record")
    optional_columns = _optional_columns(
        headers,
        tuple(OPTIONAL_RECORD_HEADERS),
        sheet_name="record",
    )
    optional = {
        normalized: (optional_columns[normalized][0], source, target)
        for source, target in OPTIONAL_RECORD_HEADERS.items()
        for normalized in [_normalize_text(source)]
        if normalized in optional_columns
    }

    data_point_index = required[_normalize_text("DataPoint")][0]
    cycle_index = required[_normalize_text("Cycle Index")][0]
    step_index = required[_normalize_text("Step Index")][0]
    status_index = required[_normalize_text("Step Type")][0]
    time_index, time_source = required[_normalize_text("Time(min)")]
    total_time_index, total_time_source = required[_normalize_text("Total Time(min)")]
    current_index = required[_normalize_text("Current(mA)")][0]
    voltage_index = required[_normalize_text("Voltage(V)")][0]
    charge_capacity_index = required[_normalize_text("Chg. Cap.(mAh)")][0]
    discharge_capacity_index = required[_normalize_text("DChg. Cap.(mAh)")][0]
    timestamp_index = required[_normalize_text("Date")][0]
    power_index, power_source = required[_normalize_text("Power(W)")]
    time_is_clock = _is_clock_duration_source(time_source, "Time(min)")
    total_time_is_clock = _is_clock_duration_source(total_time_source, "Total Time(min)")
    power_factor = _power_to_watts_factor(power_source)

    records: list[dict[str, object]] = []
    for row_number, values in enumerate(
        sheet.iter_rows(min_row=2, values_only=True), start=2
    ):
        if not any(value is not None and str(value).strip() for value in values):
            continue

        def elapsed_seconds(value: object, source_header: str, is_clock: bool) -> float:
            try:
                if is_clock:
                    return _clock_duration_seconds(value)
                number = float(value)
                if not math.isfinite(number):
                    raise ValueError("duration is not finite")
                return number * 60.0
            except (TypeError, ValueError, OverflowError) as exc:
                raise _invalid_record(row_number, source_header) from exc

        record: dict[str, object] = {
            "record_index": _integer(
                values[data_point_index], row_number=row_number, column="DataPoint"
            ),
            "cycle": _integer(
                values[cycle_index], row_number=row_number, column="Cycle Index"
            ),
            "step_index": _integer(
                values[step_index], row_number=row_number, column="Step Index"
            ),
            "status": _normalize_status(values[status_index], row_number=row_number),
            "time_s": elapsed_seconds(values[time_index], time_source, time_is_clock),
            "total_time_s": elapsed_seconds(
                values[total_time_index], total_time_source, total_time_is_clock
            ),
            "current_ma": _number(
                values[current_index], row_number=row_number, column="Current(mA)"
            ),
            "voltage_v": _number(
                values[voltage_index], row_number=row_number, column="Voltage(V)"
            ),
            "charge_capacity_mah": _number(
                values[charge_capacity_index],
                row_number=row_number,
                column="Chg. Cap.(mAh)",
            ),
            "discharge_capacity_mah": _number(
                values[discharge_capacity_index],
                row_number=row_number,
                column="DChg. Cap.(mAh)",
            ),
            "timestamp": _timestamp(
                values[timestamp_index], row_number=row_number, column="Date"
            ),
            "power_w": _number(
                values[power_index], row_number=row_number, column=power_source
            )
            * power_factor,
        }
        for normalized, (_index, source, target) in optional.items():
            record[target] = _optional_number(
                values[_index], row_number=row_number, column=source
            )
        records.append(record)

    if not records:
        raise InvalidNewareExcelError("Neware Excel record sheet contains no data rows.")
    return records


def _fast_header_map(columns: Iterator[object]) -> dict[str, tuple[int, str]]:
    result: dict[str, tuple[int, str]] = {}
    for index, value in enumerate(columns):
        original = "" if value is None else str(value).strip()
        normalized = _normalize_text(value)
        if not normalized:
            continue
        if normalized in result:
            raise InvalidNewareExcelError(
                "Neware Excel record sheet has an ambiguous normalized header: "
                f"{original}."
            )
        result[normalized] = (index, original)
    return result


def _fast_column(frame: pd.DataFrame, binding: tuple[int, str]) -> pd.Series:
    return frame.iloc[:, binding[0]]


def _fast_blank(series: pd.Series) -> pd.Series:
    if pd.api.types.is_numeric_dtype(series):
        return series.isna()
    return series.isna() | series.astype("string").str.strip().eq("").fillna(True)


def _fast_raise_first_bad(
    bad: pd.Series,
    row_numbers: np.ndarray,
    column: str,
) -> None:
    mask = bad.fillna(True).to_numpy(dtype=bool)
    positions = np.flatnonzero(mask)
    if len(positions):
        raise _invalid_record(int(row_numbers[positions[0]]), column)


def _fast_number_series(
    series: pd.Series,
    *,
    row_numbers: np.ndarray,
    column: str,
    optional: bool = False,
) -> np.ndarray:
    blank = _fast_blank(series)
    numbers = pd.to_numeric(series, errors="coerce")
    finite = pd.Series(
        np.isfinite(numbers.to_numpy(dtype="float64", na_value=np.nan)),
        index=series.index,
    )
    bad = (~blank & numbers.isna()) | (~blank & ~finite)
    if not optional:
        bad = bad | blank
    _fast_raise_first_bad(bad, row_numbers, column)
    values = numbers.to_numpy(dtype="float64", na_value=np.nan).copy()
    if optional:
        values[blank.to_numpy(dtype=bool, na_value=True)] = np.nan
    return values


def _fast_integer_series(
    series: pd.Series,
    *,
    row_numbers: np.ndarray,
    column: str,
) -> np.ndarray:
    values = _fast_number_series(series, row_numbers=row_numbers, column=column)
    int_info = np.iinfo(np.int64)
    bad = pd.Series(
        (values != np.floor(values))
        | (values < int_info.min)
        | (values >= 2**63),
        index=series.index,
    )
    _fast_raise_first_bad(bad, row_numbers, column)
    return values.astype("int64")


def _fast_duration_series(
    series: pd.Series,
    *,
    row_numbers: np.ndarray,
    source_header: str,
    canonical_header: str,
) -> np.ndarray:
    if not _is_clock_duration_source(source_header, canonical_header):
        return _fast_number_series(
            series,
            row_numbers=row_numbers,
            column=source_header,
        ) * 60.0

    values = series.to_numpy(dtype=object, na_value=None)
    seconds = np.empty(len(values), dtype="float64")
    for position, value in enumerate(values):
        if value is None or bool(pd.isna(value)):
            raise _invalid_record(int(row_numbers[position]), source_header)
        if not isinstance(value, str):
            # fastexcel normally returns clock durations as strings.  Keep the
            # reference path for native timedelta/time cells instead of
            # changing their conversion semantics in the optimized path.
            raise _FastExcelFallback("clock duration representation is not textual")
        match = _DURATION_RE.fullmatch(value.strip())
        if match is None:
            raise _invalid_record(int(row_numbers[position]), source_header)
        fraction = match.group("fraction") or ""
        fraction_seconds = int(fraction.ljust(6, "0")) / 1_000_000.0 if fraction else 0.0
        seconds[position] = (
            int(match.group("hours")) * 3600.0
            + int(match.group("minutes")) * 60.0
            + int(match.group("seconds"))
            + fraction_seconds
        )
    return seconds


def _fast_timestamp_series(
    series: pd.Series,
    *,
    row_numbers: np.ndarray,
    column: str,
) -> pd.Series:
    blank = _fast_blank(series)
    if pd.api.types.is_numeric_dtype(series):
        _fast_raise_first_bad(~blank, row_numbers, column)

    values = series.tolist()
    numeric_mask = pd.Series(
        [
            not pd.isna(value)
            and isinstance(value, Number)
            and not isinstance(value, bool)
            for value in values
        ],
        index=series.index,
    )
    _fast_raise_first_bad(numeric_mask, row_numbers, column)

    text = series.astype("string").str.strip()
    numeric_text_mask = text.str.fullmatch(_PLAIN_NUMBER_RE.pattern, na=False)
    _fast_raise_first_bad(numeric_text_mask, row_numbers, column)
    timezone_mask = text.str.contains(
        r"(?:Z|[+-]\d{2}:?\d{2})\s*$",
        regex=True,
        na=False,
    )
    _fast_raise_first_bad(timezone_mask, row_numbers, column)
    try:
        timestamps = pd.to_datetime(text, errors="coerce", format="mixed")
    except (TypeError, ValueError) as exc:
        raise _FastExcelFallback("timestamp representation is not vectorizable") from exc
    if isinstance(timestamps.dtype, pd.DatetimeTZDtype):
        _fast_raise_first_bad(~blank, row_numbers, column)
    bad = blank | timestamps.isna()
    _fast_raise_first_bad(bad, row_numbers, column)
    return pd.Series(timestamps, index=series.index, dtype="datetime64[ns]")


def _fast_status_series(
    series: pd.Series,
    *,
    row_numbers: np.ndarray,
) -> pd.Series:
    normalized = (
        series.astype("string")
        .str.replace(r"\s+", " ", regex=True)
        .str.strip()
        .str.casefold()
    )
    canonical = normalized.map(_STATUS_ALIASES)
    _fast_raise_first_bad(canonical.isna(), row_numbers, "Step Type")
    return canonical.astype("string")


def _fast_nonempty_rows(frame: pd.DataFrame) -> pd.Series:
    nonempty = pd.Series(False, index=frame.index)
    for column in frame.columns:
        series = frame[column]
        if pd.api.types.is_numeric_dtype(series):
            nonempty = nonempty | series.notna()
        else:
            nonempty = nonempty | series.astype("string").str.strip().ne("").fillna(False)
    return nonempty


class _FastFrameSheetAdapter:
    """Expose a small fastexcel frame through the reference row interface."""

    def __init__(self, title: str, frame: pd.DataFrame, *, header_row: bool):
        self.title = title
        self._frame = frame
        self._header_row = header_row

    @property
    def frame(self) -> pd.DataFrame:
        return self._frame

    @staticmethod
    def _cell(value: object) -> object:
        if value is None:
            return None
        missing = pd.isna(value)
        if isinstance(missing, (bool, np.bool_)) and bool(missing):
            return None
        return value

    def reset_dimensions(self) -> None:
        return None

    def iter_rows(
        self,
        *,
        min_row: int = 1,
        max_row: int | None = None,
        values_only: bool = True,
    ) -> Iterator[tuple[object, ...]]:
        if not values_only:
            raise ValueError("The Neware Excel parser requires values_only rows.")
        rows: Iterator[tuple[object, ...]] = (
            tuple(self._cell(value) for value in row)
            for row in self._frame.itertuples(index=False, name=None)
        )
        if self._header_row:
            rows = chain(
                (tuple(str(column) for column in self._frame.columns),),
                rows,
            )
        if min_row > 1:
            rows = islice(rows, min_row - 1, None)
        if max_row is not None:
            rows = islice(rows, max(0, max_row - min_row + 1))
        yield from rows


def _fast_load_sheet(
    reader: Any,
    name: str,
    *,
    required: bool,
    header_row: int | None,
) -> _FastFrameSheetAdapter | None:
    wanted = _normalize_text(name)
    names = [str(sheet_name) for sheet_name in reader.sheet_names]
    matches = [sheet_name for sheet_name in names if _normalize_text(sheet_name) == wanted]
    if len(matches) > 1:
        raise InvalidNewareExcelError(
            f"Neware Excel workbook has ambiguous worksheet name: {name}."
        )
    if not matches:
        if required:
            raise UnsupportedNewareExcelError(
                f"Not a recognized Neware Excel export: required {name} sheet is missing."
            )
        return None

    try:
        sheet = reader.load_sheet(
            matches[0],
            header_row=header_row,
            schema_sample_rows=1000,
            dtype_coercion="strict",
        )
        frame = sheet.to_pandas()
    except _FAST_EXCEL_ERRORS as exc:
        raise _FastExcelFallback(f"fastexcel could not represent the {name} sheet") from exc
    return _FastFrameSheetAdapter(name, frame, header_row=header_row is not None)


def _calamine_load_sheet(
    workbook: pd.ExcelFile,
    name: str,
    *,
    required: bool,
    header_row: int | None,
) -> _FastFrameSheetAdapter | None:
    """Load one sheet through pandas' native calamine adapter."""

    wanted = _normalize_text(name)
    names = [str(sheet_name) for sheet_name in workbook.sheet_names]
    matches = [sheet_name for sheet_name in names if _normalize_text(sheet_name) == wanted]
    if len(matches) > 1:
        raise InvalidNewareExcelError(
            f"Neware Excel workbook has ambiguous worksheet name: {name}."
        )
    if not matches:
        if required:
            raise UnsupportedNewareExcelError(
                f"Not a recognized Neware Excel export: required {name} sheet is missing."
            )
        return None

    try:
        frame = workbook.parse(
            matches[0],
            header=header_row,
            dtype=object,
            keep_default_na=False,
        )
    except _CALAMINE_ERRORS as exc:
        raise _CalamineFallback(
            f"pandas calamine could not represent the {name} sheet"
        ) from exc
    return _FastFrameSheetAdapter(name, frame, header_row=header_row is not None)


def _fast_declared_record_interval_seconds(reader: Any) -> float | None:
    test_sheet = _fast_load_sheet(
        reader,
        "test",
        required=False,
        header_row=None,
    )
    if test_sheet is None:
        return None
    try:
        info, _header_row, _headers = _parse_test_information(_rows(test_sheet))
    except NewareExcelError:
        return None
    settings = info.get("record_settings")
    if not isinstance(settings, dict):
        return None
    interval_s = settings.get("interval_s")
    return float(interval_s) if interval_s is not None else None


def _parse_columnar_records(
    source: pd.DataFrame,
) -> tuple[pd.DataFrame, dict[str, tuple[int, str]], bool]:
    headers = _fast_header_map(iter(source.columns))
    required = _require_columns(headers, REQUIRED_RECORD_HEADERS, sheet_name="record")
    optional_columns = _optional_columns(
        headers,
        tuple(OPTIONAL_RECORD_HEADERS),
        sheet_name="record",
    )
    nonempty = _fast_nonempty_rows(source)
    if not bool(nonempty.any()):
        raise InvalidNewareExcelError("Neware Excel record sheet contains no data rows.")
    row_numbers = source.index[nonempty].to_numpy(dtype="int64") + 2
    source = source.loc[nonempty].reset_index(drop=True)

    record_index = _fast_integer_series(
        _fast_column(source, required[_normalize_text("DataPoint")]),
        row_numbers=row_numbers,
        column="DataPoint",
    )
    cycle = _fast_integer_series(
        _fast_column(source, required[_normalize_text("Cycle Index")]),
        row_numbers=row_numbers,
        column="Cycle Index",
    )
    step_index = _fast_integer_series(
        _fast_column(source, required[_normalize_text("Step Index")]),
        row_numbers=row_numbers,
        column="Step Index",
    )
    time_binding = required[_normalize_text("Time(min)")]
    total_time_binding = required[_normalize_text("Total Time(min)")]
    time_s = _fast_duration_series(
        _fast_column(source, time_binding),
        row_numbers=row_numbers,
        source_header=time_binding[1],
        canonical_header="Time(min)",
    )
    total_time_s = _fast_duration_series(
        _fast_column(source, total_time_binding),
        row_numbers=row_numbers,
        source_header=total_time_binding[1],
        canonical_header="Total Time(min)",
    )
    data: dict[str, Any] = {
        "record_index": record_index,
        "cycle": cycle,
        "step_index": step_index,
        "status": _fast_status_series(
            _fast_column(source, required[_normalize_text("Step Type")]),
            row_numbers=row_numbers,
        ),
        "time_s": time_s,
        "total_time_s": total_time_s,
        "current_ma": _fast_number_series(
            _fast_column(source, required[_normalize_text("Current(mA)")]),
            row_numbers=row_numbers,
            column="Current(mA)",
        ),
        "voltage_v": _fast_number_series(
            _fast_column(source, required[_normalize_text("Voltage(V)")]),
            row_numbers=row_numbers,
            column="Voltage(V)",
        ),
        "charge_capacity_mah": _fast_number_series(
            _fast_column(source, required[_normalize_text("Chg. Cap.(mAh)")]),
            row_numbers=row_numbers,
            column="Chg. Cap.(mAh)",
        ),
        "discharge_capacity_mah": _fast_number_series(
            _fast_column(source, required[_normalize_text("DChg. Cap.(mAh)")]),
            row_numbers=row_numbers,
            column="DChg. Cap.(mAh)",
        ),
        "timestamp": _fast_timestamp_series(
            _fast_column(source, required[_normalize_text("Date")]),
            row_numbers=row_numbers,
            column="Date",
        ),
    }
    power_binding = required[_normalize_text("Power(W)")]
    data["power_w"] = _fast_number_series(
        _fast_column(source, power_binding),
        row_numbers=row_numbers,
        column=power_binding[1],
    ) * _power_to_watts_factor(power_binding[1])
    for canonical, target in OPTIONAL_RECORD_HEADERS.items():
        binding = optional_columns.get(_normalize_text(canonical))
        if binding is None:
            continue
        data[target] = _fast_number_series(
            _fast_column(source, binding),
            row_numbers=row_numbers,
            column=binding[1],
            optional=True,
        )

    order = np.argsort(record_index, kind="stable")
    ordered_index = record_index[order]
    duplicate = np.zeros(len(ordered_index), dtype=bool)
    if len(ordered_index) > 1:
        duplicate[1:] = ordered_index[1:] == ordered_index[:-1]
    _fast_raise_first_bad(
        pd.Series(duplicate),
        row_numbers[order],
        "DataPoint",
    )
    for key, values in tuple(data.items()):
        if isinstance(values, pd.Series):
            data[key] = values.iloc[order].reset_index(drop=True)
        else:
            data[key] = np.asarray(values)[order]

    frame = _frame_from_data(data)
    record_clock_dialect = (
        _is_clock_duration_source(time_binding[1], "Time(min)")
        and _is_clock_duration_source(total_time_binding[1], "Total Time(min)")
    )
    return frame, headers, record_clock_dialect


def _parse_fast_records(
    path: Path,
) -> tuple[pd.DataFrame, dict[str, tuple[int, str]], bool, Any]:
    if _fastexcel is None:
        raise _FastExcelFallback("fastexcel is not installed")

    try:
        reader = _fastexcel.read_excel(path)
        record_sheet = _fast_load_sheet(
            reader,
            "record",
            required=True,
            header_row=0,
        )
        if record_sheet is None:  # pragma: no cover - required=True raises above
            raise _FastExcelFallback("fastexcel did not return the record sheet")
    except _FAST_EXCEL_ERRORS as exc:
        raise _FastExcelFallback("fastexcel could not represent the record sheet") from exc

    frame, headers, record_clock_dialect = _parse_columnar_records(record_sheet.frame)
    return frame, headers, record_clock_dialect, reader


def _parse_fast_timeseries(
    path: Path,
) -> tuple[pd.DataFrame, dict[str, tuple[int, str]], bool, bool] | None:
    try:
        frame, record_headers, record_clock_dialect, reader = _parse_fast_records(path)
        step_sheet = _fast_load_sheet(
            reader,
            "step",
            required=False,
            header_row=0,
        )
        step_duration_validated = False
        if step_sheet is not None:
            step_duration_validated = _validate_step_summary(
                frame,
                step_sheet,
                record_interval_s=_fast_declared_record_interval_seconds(reader),
                record_clock_dialect=record_clock_dialect,
            )
        return frame, record_headers, record_clock_dialect, step_duration_validated
    except _FastExcelFallback:
        return None


def _parse_calamine_timeseries(
    path: Path,
) -> tuple[pd.DataFrame, dict[str, tuple[int, str]], bool, bool] | None:
    """Parse through pandas' calamine engine before using openpyxl."""

    if _python_calamine is None:
        return None

    workbook: pd.ExcelFile | None = None
    try:
        try:
            workbook = pd.ExcelFile(path, engine="calamine")
        except _CALAMINE_ERRORS:
            return None

        record_sheet = _calamine_load_sheet(
            workbook,
            "record",
            required=True,
            header_row=0,
        )
        if record_sheet is None:  # pragma: no cover - required=True raises above
            raise _CalamineFallback("pandas calamine did not return the record sheet")
        frame, record_headers, record_clock_dialect = _parse_columnar_records(
            record_sheet.frame
        )
        step_sheet = _calamine_load_sheet(
            workbook,
            "step",
            required=False,
            header_row=0,
        )
        step_duration_validated = False
        if step_sheet is not None:
            test_sheet = _calamine_load_sheet(
                workbook,
                "test",
                required=False,
                header_row=None,
            )
            record_interval_s: float | None = None
            if test_sheet is not None:
                try:
                    info, _header_row, _headers = _parse_test_information(_rows(test_sheet))
                except NewareExcelError:
                    info = {}
                settings = info.get("record_settings")
                if isinstance(settings, dict) and settings.get("interval_s") is not None:
                    record_interval_s = float(settings["interval_s"])
            step_duration_validated = _validate_step_summary(
                frame,
                step_sheet,
                record_interval_s=record_interval_s,
                record_clock_dialect=record_clock_dialect,
            )
        return frame, record_headers, record_clock_dialect, step_duration_validated
    except (_CalamineFallback, _FastExcelFallback):
        return None
    finally:
        if workbook is not None:
            workbook.close()


def _ordered_records(records: list[dict[str, object]]) -> list[dict[str, object]]:
    order = sorted(range(len(records)), key=lambda index: int(records[index]["record_index"]))
    ordered = [records[index] for index in order]
    previous: int | None = None
    for record in ordered:
        current = int(record["record_index"])
        if previous is not None and current <= previous:
            if current == previous:
                raise InvalidNewareExcelError(
                    f"Neware Excel record DataPoint {current} is duplicated."
                )
            raise InvalidNewareExcelError(
                "Neware Excel DataPoint values must be strictly increasing."
            )
        previous = current
    return ordered


def _frame_from_records(records: list[dict[str, object]]) -> pd.DataFrame:
    ordered = _ordered_records(records)
    data: dict[str, list[object]] = {}
    for key in ordered[0]:
        data[key] = [record[key] for record in ordered]

    return _frame_from_data(data)


def _frame_from_data(data: dict[str, Any]) -> pd.DataFrame:
    """Build the canonical frame and derived fields from ordered columns."""

    frame = pd.DataFrame(data)
    frame["record_index"] = pd.Series(data["record_index"], dtype="int64")
    frame["cycle"] = pd.Series(data["cycle"], dtype="int64")
    frame["step_index"] = pd.Series(data["step_index"], dtype="int64")
    for column in _FLOAT_COLUMNS:
        if column in frame:
            frame[column] = pd.Series(data[column], dtype="float64")
    frame["status"] = pd.Series(data["status"], dtype="string")
    frame["timestamp"] = pd.Series(data["timestamp"], dtype="datetime64[ns]")

    total_time = frame["total_time_s"].to_numpy(dtype="float64")
    if len(total_time) > 1 and np.any(np.diff(total_time) < -1e-9):
        raise InvalidNewareExcelError(
            "Neware Excel record Total Time(min) decreases in source order."
        )

    cycle = frame["cycle"].to_numpy(dtype="int64")
    step_index = frame["step_index"].to_numpy(dtype="int64")
    time_s = frame["time_s"].to_numpy(dtype="float64")
    status = frame["status"].astype(str).to_numpy()
    executed_steps = np.empty(len(frame), dtype="int64")
    step_number = 0
    for index in range(len(frame)):
        boundary = (
            index == 0
            or cycle[index] != cycle[index - 1]
            or step_index[index] != step_index[index - 1]
            or status[index] != status[index - 1]
            or time_s[index] < time_s[index - 1] - 1e-9
        )
        if boundary:
            step_number += 1
        executed_steps[index] = step_number
    frame["step"] = executed_steps

    charge_energy = np.zeros(len(frame), dtype="float64")
    discharge_energy = np.zeros(len(frame), dtype="float64")
    power = frame["power_w"].to_numpy(dtype="float64")
    for index in range(1, len(frame)):
        if executed_steps[index] != executed_steps[index - 1]:
            continue
        dt_h = max(0.0, total_time[index] - total_time[index - 1]) / 3600.0
        increment = abs((power[index - 1] + power[index]) / 2.0) * dt_h * 1000.0
        if "DChg" in status[index]:
            discharge_energy[index] = discharge_energy[index - 1] + increment
        elif "Chg" in status[index] and "DChg" not in status[index]:
            charge_energy[index] = charge_energy[index - 1] + increment

    frame["charge_energy_mwh"] = charge_energy
    frame["discharge_energy_mwh"] = discharge_energy
    columns = [
        "record_index",
        "cycle",
        "step",
        "step_index",
        "status",
        "time_s",
        "total_time_s",
        "voltage_v",
        "current_ma",
        "charge_capacity_mah",
        "discharge_capacity_mah",
        "charge_energy_mwh",
        "discharge_energy_mwh",
        "timestamp",
        "power_w",
    ]
    columns.extend(
        target for target in OPTIONAL_RECORD_HEADERS.values() if target in frame.columns
    )
    return frame[columns]


def _parse_step_summary(sheet: Any) -> list[dict[str, object]]:
    headers = _header_map(sheet)
    required = _require_columns(headers, STEP_HEADERS, sheet_name="step")
    step_time_source = required[_normalize_text("Step Time(min)")][1]
    energy_source = required[_normalize_text("Energy(Wh)")][1]
    rows: list[dict[str, object]] = []
    for row_number, values in enumerate(
        sheet.iter_rows(min_row=2, values_only=True), start=2
    ):
        if not any(value is not None and str(value).strip() for value in values):
            continue

        def value_for(source: str) -> object:
            return values[required[_normalize_text(source)][0]]

        def summary_number(source: str) -> float:
            value = value_for(source)
            if value is None or (isinstance(value, str) and not value.strip()):
                raise InvalidNewareExcelError(
                    f"Neware Excel step summary row {row_number} has an invalid {source} value."
                )
            try:
                number = float(value)
            except (TypeError, ValueError) as exc:
                raise InvalidNewareExcelError(
                    f"Neware Excel step summary row {row_number} has an invalid {source} value."
                ) from exc
            if not math.isfinite(number):
                raise InvalidNewareExcelError(
                    f"Neware Excel step summary row {row_number} has an invalid {source} value."
                )
            return number

        def summary_integer(source: str) -> int:
            number = summary_number(source)
            if not number.is_integer():
                raise InvalidNewareExcelError(
                    f"Neware Excel step summary row {row_number} has an invalid {source} value."
                )
            return int(number)

        def summary_timestamp(source: str) -> pd.Timestamp:
            try:
                timestamp = _coerce_timestamp(value_for(source))
            except (TypeError, ValueError, OverflowError) as exc:
                raise InvalidNewareExcelError(
                    f"Neware Excel step summary row {row_number} has an invalid {source} value."
                ) from exc
            return timestamp

        def summary_elapsed(source: str, source_header: str) -> float:
            value = value_for(source)
            try:
                return _unitless_or_minutes_seconds(
                    value,
                    source_header=source_header,
                    canonical_header=source,
                )
            except (TypeError, ValueError, OverflowError) as exc:
                raise InvalidNewareExcelError(
                    f"Neware Excel step summary row {row_number} has an invalid {source_header} value."
                ) from exc

        energy_rounding_tolerance = 0.01

        def summary_energy(source: str, source_header: str) -> float:
            nonlocal energy_rounding_tolerance
            source_number = summary_number(source)
            factor = _energy_to_mwh_factor(source_header)
            energy_rounding_tolerance = _display_rounding_tolerance(
                source_number,
                scale=factor,
                minimum=0.01,
            )
            return source_number * factor

        rows.append(
            {
                "cycle": summary_integer("Cycle Index"),
                "step_index": summary_integer("Step Index"),
                "step_number": summary_integer("Step Number"),
                "status": _normalize_status(
                    value_for("Step Type"), row_number=row_number, column="Step Type"
                ),
                "step_time_s": summary_elapsed("Step Time(min)", step_time_source),
                "onset": summary_timestamp("Oneset Date"),
                "end": summary_timestamp("End Date"),
                "capacity_mah": summary_number("Capacity(mAh)"),
                "energy_mwh": summary_energy("Energy(Wh)", energy_source),
                "energy_rounding_tolerance_mwh": energy_rounding_tolerance,
                "onset_voltage_v": summary_number("Oneset Volt.(V)"),
                "end_voltage_v": summary_number("End Voltage(V)"),
            }
        )

    if not rows:
        raise InvalidNewareExcelError("Neware Excel step sheet contains no data rows.")
    rows.sort(key=lambda row: int(row["step_number"]))
    previous: int | None = None
    for row in rows:
        current = int(row["step_number"])
        if previous is not None and current <= previous:
            raise InvalidNewareExcelError(
                "Neware Excel step summary Step Number values must be unique and increasing."
            )
        previous = current
    return rows


def _segment_groups(frame: pd.DataFrame) -> list[pd.DataFrame]:
    step_values = frame["step"].to_numpy(dtype="int64")
    starts = np.flatnonzero(
        np.r_[True, step_values[1:] != step_values[:-1]]
    )
    ends = np.r_[starts[1:], len(frame)]
    return [frame.iloc[start:end] for start, end in zip(starts, ends)]


def _segment_capacity(group: pd.DataFrame) -> float:
    status = str(group["status"].iloc[0])
    column = "discharge_capacity_mah" if "DChg" in status else "charge_capacity_mah" if "Chg" in status else None
    if column is None:
        return 0.0
    values = group[column].to_numpy(dtype="float64")
    return float(np.max(values) - np.min(values))


def _integrate_step_energy(group: pd.DataFrame) -> float:
    total_time = group["total_time_s"].to_numpy(dtype="float64")
    power = group["power_w"].to_numpy(dtype="float64")
    if len(group) < 2:
        return 0.0
    energy = 0.0
    for index in range(1, len(group)):
        dt_h = max(0.0, total_time[index] - total_time[index - 1]) / 3600.0
        energy += abs((power[index - 1] + power[index]) / 2.0) * dt_h * 1000.0
    return energy


def _step_time_tolerance_seconds(record_interval_s: float | None) -> float:
    """Return the locked declared-cadence timing tolerance."""

    return max(2.0, float(record_interval_s)) if record_interval_s is not None else 2.0


def _validate_step_summary(
    frame: pd.DataFrame,
    sheet: Any,
    *,
    record_interval_s: float | None,
    record_clock_dialect: bool,
) -> bool:
    summary = _parse_step_summary(sheet)
    step_time_source = _require_columns(
        _header_map(sheet),
        STEP_HEADERS,
        sheet_name="step",
    )[_normalize_text("Step Time(min)")][1]
    duration_is_clock = _is_clock_duration_source(
        step_time_source,
        "Step Time(min)",
    )
    relaxed_clock_dialect = duration_is_clock and record_clock_dialect
    segments = _segment_groups(frame)
    if len(summary) != len(segments):
        raise InvalidNewareExcelError(
            "Neware Excel execution-step mapping failed: "
            f"{len(segments)} raw segments but {len(summary)} step-summary rows."
        )

    for expected_step, (summary_row, segment) in enumerate(
        zip(summary, segments), start=1
    ):
        actual_step = int(segment["step"].iloc[0])
        if int(summary_row["step_number"]) != actual_step:
            raise InvalidNewareExcelError(
                "Neware Excel execution-step mapping failed: Step Number does not match raw order."
            )
        identity = (
            int(segment["cycle"].iloc[0]) == int(summary_row["cycle"])
            and int(segment["step_index"].iloc[0]) == int(summary_row["step_index"])
            and str(segment["status"].iloc[0]) == str(summary_row["status"])
        )
        if not identity:
            raise InvalidNewareExcelError(
                f"Neware Excel execution-step mapping failed at step {expected_step}."
            )

        onset = pd.Timestamp(segment["timestamp"].iloc[0])
        end = pd.Timestamp(segment["timestamp"].iloc[-1])
        tolerance_s = _step_time_tolerance_seconds(record_interval_s)
        if abs((onset - summary_row["onset"]).total_seconds()) > tolerance_s:
            raise InvalidNewareExcelError(
                f"Neware Excel step summary onset does not match raw step {expected_step}."
            )
        if abs((end - summary_row["end"]).total_seconds()) > tolerance_s:
            raise InvalidNewareExcelError(
                f"Neware Excel step summary end does not match raw step {expected_step}."
            )
        if duration_is_clock:
            elapsed_values = segment["time_s"].to_numpy(dtype="float64")
            # The unitless clock dialect is an execution-relative duration.
            # Its timestamp span can include a paused/restored gap, so compare
            # the summary with the record's step-relative elapsed-time column.
            duration_s = float(np.max(elapsed_values) - np.min(elapsed_values))
        else:
            # Preserve the original numeric-dialect contract: Step Time(min)
            # is reconciled independently with the exported timestamps.
            duration_s = (end - onset).total_seconds()
        if abs(duration_s - float(summary_row["step_time_s"])) > tolerance_s:
            raise InvalidNewareExcelError(
                f"Neware Excel step summary duration does not match raw step {expected_step}."
            )

        start_voltage = float(segment["voltage_v"].iloc[0])
        end_voltage = float(segment["voltage_v"].iloc[-1])
        if abs(start_voltage - float(summary_row["onset_voltage_v"])) > 0.002:
            raise InvalidNewareExcelError(
                f"Neware Excel step summary onset voltage does not match raw step {expected_step}."
            )
        if abs(end_voltage - float(summary_row["end_voltage_v"])) > 0.002:
            raise InvalidNewareExcelError(
                f"Neware Excel step summary end voltage does not match raw step {expected_step}."
            )

        expected_capacity = float(summary_row["capacity_mah"])
        capacity_tolerance = max(0.002, abs(expected_capacity) * 0.001)
        if relaxed_clock_dialect:
            capacity_tolerance = max(
                capacity_tolerance,
                _display_rounding_tolerance(expected_capacity, scale=1.0, minimum=0.0) * 2.5,
            )
        if abs(_segment_capacity(segment) - expected_capacity) > capacity_tolerance:
            raise InvalidNewareExcelError(
                f"Neware Excel step summary capacity does not match raw step {expected_step}."
            )

        expected_energy = float(summary_row["energy_mwh"])
        energy_tolerance = max(0.01, abs(expected_energy) * 0.001)
        if relaxed_clock_dialect:
            energy_tolerance = max(
                energy_tolerance,
                float(summary_row.get("energy_rounding_tolerance_mwh", 0.01)) * 2.5,
                abs(expected_energy) * 0.005,
            )
        if abs(_integrate_step_energy(segment) - expected_energy) > energy_tolerance:
            raise InvalidNewareExcelError(
                f"Neware Excel step summary energy does not match raw step {expected_step}."
            )
    return True


def is_supported_workbook(path: str | Path) -> bool:
    """Return whether ``path`` matches the supported Neware record contract."""

    candidate = _path(path)
    if candidate.suffix.casefold() != ".xlsx":
        return False
    try:
        with _open(candidate) as workbook:
            sheet = _sheet_by_name(workbook, "record", required=True)
            headers = _header_map(sheet)
            _require_columns(headers, REQUIRED_RECORD_HEADERS, sheet_name="record")
        return True
    except NewareExcelError:
        return False


def parse_timeseries(path: str | Path) -> pd.DataFrame:
    """Parse a supported Neware Excel export into canonical raw data."""

    candidate = _path(path)
    if candidate.suffix.casefold() != ".xlsx":
        raise UnsupportedNewareExcelError(
            "Not a recognized Neware Excel export: only .xlsx is supported."
        )

    try:
        fast_result = _parse_fast_timeseries(candidate)

        if fast_result is not None:
            frame, record_headers, record_clock_dialect, step_duration_validated = fast_result
            step_sheet = True if step_duration_validated else None
        else:
            calamine_result = _parse_calamine_timeseries(candidate)
            if calamine_result is not None:
                frame, record_headers, record_clock_dialect, step_duration_validated = (
                    calamine_result
                )
                step_sheet = True if step_duration_validated else None
            else:
                with _open(candidate) as workbook:
                    record_sheet = _sheet_by_name(workbook, "record", required=True)
                    record_headers = _header_map(record_sheet)
                    records = _parse_records(record_sheet, record_headers)
                    frame = _frame_from_records(records)
                    record_time_source = _require_columns(
                        record_headers,
                        REQUIRED_RECORD_HEADERS,
                        sheet_name="record",
                    )[_normalize_text("Time(min)")][1]
                    record_total_time_source = _require_columns(
                        record_headers,
                        REQUIRED_RECORD_HEADERS,
                        sheet_name="record",
                    )[_normalize_text("Total Time(min)")][1]
                    record_clock_dialect = (
                        _is_clock_duration_source(record_time_source, "Time(min)")
                        and _is_clock_duration_source(record_total_time_source, "Total Time(min)")
                    )
                    step_sheet = _sheet_by_name(workbook, "step", required=False)
                    step_duration_validated = False
                    if step_sheet is not None:
                        step_duration_validated = _validate_step_summary(
                            frame,
                            step_sheet,
                            record_interval_s=_declared_record_interval_seconds(workbook),
                            record_clock_dialect=record_clock_dialect,
                        )
    except NewareExcelError:
        raise
    except Exception as exc:
        raise InvalidNewareExcelError(
            "Could not parse the Neware Excel workbook."
        ) from exc

    frame.attrs["neware_excel"] = {
        "record_sheet": "record",
        "step_summary_available": step_sheet is not None,
        "step_summary_validated": step_sheet is not None,
        "step_summary_duration_validated": step_duration_validated,
        "record_clock_dialect": record_clock_dialect,
        "record_count": int(len(frame)),
        "executed_step_count": int(frame["step"].nunique()),
    }
    return frame


def read_metadata(path: str | Path) -> dict[str, object]:
    """Read bounded workbook metadata and the programmed plan.

    The metadata path intentionally opens only the small ``test``, ``unit`` and
    sheet/header surfaces.  It never iterates the large ``record`` worksheet and
    does not derive metadata from :func:`parse_timeseries`.
    """

    candidate = _path(path)
    if candidate.suffix.casefold() != ".xlsx":
        raise UnsupportedNewareExcelError(
            "Not a recognized Neware Excel export: only .xlsx is supported."
        )

    try:
        with _open(candidate) as workbook:
            record_sheet = _sheet_by_name(workbook, "record", required=True)
            _require_columns(
                _header_map(record_sheet),
                REQUIRED_RECORD_HEADERS,
                sheet_name="record",
            )
            test_sheet = _sheet_by_name(workbook, "test", required=False)
            info: dict[str, object] = {"raw": {}}
            step_info: dict[str, dict[str, str]] = {}
            original_steps: dict[str, dict[str, object]] = {}
            if test_sheet is not None:
                test_rows = _rows(test_sheet)
                info, header_row, headers = _parse_test_information(test_rows)
                step_info, original_steps = _parse_programmed_plan(
                    test_rows, header_row, headers, info
                )
            unit_original, unit_workbook_name = _parse_unit_original(
                _sheet_by_name(workbook, "unit", required=False)
            )
            has_cycle_summary = _sheet_by_name(workbook, "cycle", required=False) is not None
            has_step_summary = _sheet_by_name(workbook, "step", required=False) is not None
    except NewareExcelError:
        raise
    except Exception as exc:
        raise InvalidNewareExcelError(
            "Could not read Neware Excel metadata."
        ) from exc

    if info.get("start_time") is None:
        start_time = unit_original.get("StartTime", {}).get("Value")
        if start_time:
            info["start_time"] = start_time

    head_info: dict[str, object] = {}

    def head_value(key: str, value: object) -> None:
        if value is not None and not _is_blank(value):
            head_info[key] = {"Value": str(value)}

    head_value("Start_Step", info.get("start_step_id"))
    head_value("PN", info.get("part_number"))
    head_value("Creator", info.get("builder"))
    head_value("Remark", info.get("remarks"))
    if info.get("active_mass_mg") is not None:
        head_value("SCQ", float(info["active_mass_mg"]) * 1000.0)
    if info.get("nominal_capacity_mah") is not None:
        head_value("MultCap", float(info["nominal_capacity_mah"]) * 3600.0)

    protection: dict[str, object] = {}
    if info.get("protection_voltage_upper_v") is not None:
        protection["Upper"] = {
            "Value": _format_number(float(info["protection_voltage_upper_v"]) * 10000.0)
        }
    if info.get("protection_voltage_lower_v") is not None:
        protection["Lower"] = {
            "Value": _format_number(float(info["protection_voltage_lower_v"]) * 10000.0)
        }
    if protection:
        head_info["Protect"] = {"Main": {"Volt": protection}}

    original_test: dict[str, object] = {}
    original_names = {
        "start_step_id": "StartStepID",
        "protection_voltage_upper_v": "VoltUpper",
        "protection_voltage_lower_v": "VoltLower",
        "builder": "Builder",
        "remarks": "Remarks",
        "start_time": "StartTime",
        "barcode": "Barcode",
        "active_mass_mg": "ActiveMaterial",
        "nominal_capacity_mah": "NominalCapacity",
        "part_number": "PartNumber",
        "cycle_count": "CycleCount",
        "voltage_range": "VoltageRange",
        "current_range": "CurrentRange",
    }
    raw_info = info["raw"]
    for source_key, output_key in original_names.items():
        raw_value = raw_info.get(source_key)
        if not _is_blank(raw_value):
            original_test[output_key] = {"Value": _value_text(raw_value)}
    record_settings = info.get("record_settings")
    if record_settings is not None:
        original_test["RecordSettings"] = {
            "Value": str(record_settings["raw"]),
            "IntervalS": {"Value": _format_number(float(record_settings["interval_s"]))},
            "VoltageDeltaV": {"Value": _format_number(float(record_settings["voltage_delta_v"]))},
            "CurrentDeltaMA": {"Value": _format_number(float(record_settings["current_delta_ma"]))},
        }
    if info.get("start_time") is not None and "StartTime" not in original_test:
        original_test["StartTime"] = {"Value": str(info["start_time"])}

    original: dict[str, object] = {"Test": original_test, "StepPlan": original_steps}
    if unit_original:
        original["Unit"] = unit_original
    if unit_workbook_name and "WorkbookName" not in unit_original:
        original.setdefault("Unit", {})["WorkbookName"] = {"Value": unit_workbook_name}

    return {
        "Step": {
            "Head_Info": head_info,
            "Step_Info": step_info,
        },
        "Excel": {
            "SourceFormat": {"Value": "neware_excel"},
            "ParserRevision": {"Value": str(EXCEL_PARSER_REVISION)},
            "Capabilities": {
                "ExecutedStepSummary": {"Value": has_step_summary},
                "CycleSummary": {"Value": has_cycle_summary},
                "DeclaredProtocol": {"Value": test_sheet is not None},
                "ProtocolConditions": {"Value": False},
            },
            "Original": original,
        },
    }


def _summary_number(value: object, *, row_number: int, column: str) -> float | None:
    if _is_blank(value):
        return None
    try:
        number = float(value)
    except (TypeError, ValueError) as exc:
        raise InvalidNewareExcelError(
            f"Neware Excel cycle summary row {row_number} has an invalid {column} value."
        ) from exc
    if not math.isfinite(number):
        raise InvalidNewareExcelError(
            f"Neware Excel cycle summary row {row_number} has an invalid {column} value."
        )
    return number


def _display_rounding_tolerance(value: float, *, scale: float, minimum: float) -> float:
    """Account for the precision displayed by a Neware summary cell."""

    try:
        decimal_value = Decimal(str(value))
    except (InvalidOperation, ValueError):
        return minimum
    exponent = decimal_value.as_tuple().exponent
    if not isinstance(exponent, int) or exponent >= 0:
        return minimum
    return max(minimum, 0.5 * (10.0**exponent) * scale)


def _parse_cycle_summary(sheet: Any) -> list[dict[str, float | int | None]]:
    headers = _header_map(sheet)
    required = _require_columns(headers, _CYCLE_REQUIRED_HEADERS, sheet_name="cycle")
    optional = _optional_columns(
        headers,
        _CYCLE_OPTIONAL_HEADERS,
        sheet_name="cycle",
    )
    charge_time_source = required[_normalize_text("Chg. Time(min)")][1]
    discharge_time_source = required[_normalize_text("DChg. Time(min)")][1]
    charge_time_is_clock = _is_clock_duration_source(
        charge_time_source,
        "Chg. Time(min)",
    )
    discharge_time_is_clock = _is_clock_duration_source(
        discharge_time_source,
        "DChg. Time(min)",
    )
    parsed: list[dict[str, float | int | None | bool]] = []
    for row_number, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if not any(not _is_blank(value) for value in values):
            continue

        def value_for(header: str) -> object:
            binding = required.get(_normalize_text(header)) or optional.get(_normalize_text(header))
            return None if binding is None else values[binding[0]]

        energy_rounding_tolerances: dict[str, float] = {}

        def scaled_energy_value(header: str) -> float | None:
            binding = optional.get(_normalize_text(header))
            if binding is None:
                return None
            number = _summary_number(value_for(header), row_number=row_number, column=header)
            if number is None:
                return None
            factor = _energy_to_mwh_factor(binding[1])
            energy_rounding_tolerances[header] = _display_rounding_tolerance(
                number,
                scale=factor,
                minimum=0.0,
            )
            return number * factor

        def elapsed_value(header: str, source_header: str) -> float | None:
            value = value_for(header)
            if _is_blank(value):
                return None
            try:
                return _unitless_or_minutes_seconds(
                    value,
                    source_header=source_header,
                    canonical_header=header,
                )
            except (TypeError, ValueError, OverflowError) as exc:
                raise InvalidNewareExcelError(
                    f"Neware Excel cycle summary row {row_number} has an invalid {source_header} value."
                ) from exc

        cycle = _integer_value(value_for("Cycle Index"), label="Cycle Index", required=True)
        if cycle is None or cycle <= 0:
            raise InvalidNewareExcelError(
                f"Neware Excel cycle summary row {row_number} has an invalid Cycle Index."
            )
        charge_capacity = _summary_number(
            value_for("Chg. Cap.(mAh)"),
            row_number=row_number,
            column="Chg. Cap.(mAh)",
        )
        discharge_capacity = _summary_number(
            value_for("DChg. Cap.(mAh)"),
            row_number=row_number,
            column="DChg. Cap.(mAh)",
        )
        efficiency = _summary_number(
            value_for("Chg.-DChg. Eff(%)"),
            row_number=row_number,
            column="Chg.-DChg. Eff(%)",
        )
        if (
            efficiency is None
            and charge_capacity is not None
            and discharge_capacity is not None
            and not math.isclose(charge_capacity, 0.0, abs_tol=1e-12)
        ):
            efficiency = discharge_capacity / charge_capacity * 100.0

        parsed.append(
            {
                "cycle": cycle,
                "charge_capacity_mah": charge_capacity,
                "discharge_capacity_mah": discharge_capacity,
                "coulombic_efficiency_pct": efficiency,
                "charge_energy_mwh": scaled_energy_value("Chg. Energy(Wh)"),
                "discharge_energy_mwh": scaled_energy_value("DChg. Energy(Wh)"),
                "charge_energy_rounding_tolerance_mwh": energy_rounding_tolerances.get(
                    "Chg. Energy(Wh)", 0.0
                ),
                "discharge_energy_rounding_tolerance_mwh": energy_rounding_tolerances.get(
                    "DChg. Energy(Wh)", 0.0
                ),
                "charge_time_is_clock": charge_time_is_clock,
                "discharge_time_is_clock": discharge_time_is_clock,
                "charge_time_s": elapsed_value("Chg. Time(min)", charge_time_source),
                "discharge_time_s": elapsed_value("DChg. Time(min)", discharge_time_source),
            }
        )

    if not parsed:
        raise InvalidNewareExcelError("Neware Excel cycle sheet contains no data rows.")
    parsed.sort(key=lambda row: int(row["cycle"]))
    cycle_ids = [int(row["cycle"]) for row in parsed]
    if len(set(cycle_ids)) != len(cycle_ids):
        raise InvalidNewareExcelError("Neware Excel cycle summary contains duplicate Cycle Index values.")
    return parsed


def validate_cycles(path: str | Path, raw: pd.DataFrame, cycles: pd.DataFrame) -> None:
    """Cross-check calculated cycles against the workbook's small cycle sheet."""

    candidate = _path(path)
    if candidate.suffix.casefold() != ".xlsx":
        return

    state = raw.attrs.setdefault("neware_excel", {})
    with _open(candidate) as workbook:
        cycle_sheet = _sheet_by_name(workbook, "cycle", required=False)
        if cycle_sheet is None:
            state["cycle_summary_available"] = False
            state["cycle_summary_validated"] = False
            return
        summary = _parse_cycle_summary(cycle_sheet)
        interval_s: float | None = None
        test_sheet = _sheet_by_name(workbook, "test", required=False)
        if test_sheet is not None:
            try:
                info, _header_row, _headers = _parse_test_information(_rows(test_sheet))
                settings = info.get("record_settings")
                if settings is not None:
                    interval_s = float(settings["interval_s"])
            except NewareExcelError:
                # Cycle validation remains useful when optional metadata is
                # absent or malformed; the documented two-second floor applies.
                interval_s = None

    state["cycle_summary_available"] = True
    state["cycle_summary_validated"] = False
    if cycles is None or cycles.empty or "cycle" not in cycles.columns:
        raise InvalidNewareExcelError("Neware Excel cycle summary cannot be compared with empty calculated cycles.")

    actual = cycles.sort_values("cycle", kind="stable").reset_index(drop=True)
    actual_ids = [int(value) for value in actual["cycle"].tolist()]
    summary_ids = [int(row["cycle"]) for row in summary]
    if actual_ids != summary_ids:
        raise InvalidNewareExcelError(
            "Neware Excel cycle summary identity mismatch: cycle count or order differs."
        )

    time_tolerance_s = max(2.0, interval_s if interval_s is not None else 2.0)
    record_clock_dialect = bool(state.get("record_clock_dialect", False))

    def compare(
        cycle_id: int,
        quantity: str,
        actual_value: object,
        expected_value: object,
        tolerance: float,
    ) -> None:
        if expected_value is None:
            return
        duration_quantity = quantity in {"charge time", "discharge time"}
        try:
            actual_number = float(actual_value)
            expected_number = float(expected_value)
        except (TypeError, ValueError) as exc:
            try:
                expected_number = float(expected_value)
            except (TypeError, ValueError):
                raise InvalidNewareExcelError(
                    f"Neware Excel cycle {cycle_id} {quantity} cannot be compared."
                ) from exc
            # Neware omits a derived duration when a cycle has no matching
            # phase, while its summary writes the corresponding value as 0.
            # Treat that representation as the same zero rather than turning
            # an otherwise valid final cycle into an import failure.
            if duration_quantity and abs(expected_number) <= tolerance and (
                actual_value is None or pd.isna(actual_value)
            ):
                return
            raise InvalidNewareExcelError(
                f"Neware Excel cycle {cycle_id} {quantity} cannot be compared."
            ) from exc
        if not math.isfinite(expected_number):
            raise InvalidNewareExcelError(
                f"Neware Excel cycle {cycle_id} {quantity} has a non-finite summary value."
            )
        if math.isnan(actual_number):
            if duration_quantity and abs(expected_number) <= tolerance:
                return
            raise InvalidNewareExcelError(
                f"Neware Excel cycle {cycle_id} {quantity} mismatch: calculated {actual_number:g}, "
                f"summary {expected_number:g}."
            )
        if not math.isfinite(actual_number):
            raise InvalidNewareExcelError(
                f"Neware Excel cycle {cycle_id} {quantity} mismatch: calculated {actual_number:g}, "
                f"summary {expected_number:g}."
            )
        if abs(actual_number - expected_number) > tolerance:
            raise InvalidNewareExcelError(
                f"Neware Excel cycle {cycle_id} {quantity} mismatch: "
                f"calculated {actual_number:g}, summary {expected_number:g}."
            )

    for index, row in enumerate(summary):
        cycle_id = int(row["cycle"])
        actual_row = actual.iloc[index]
        charge_capacity = row["charge_capacity_mah"]
        discharge_capacity = row["discharge_capacity_mah"]
        charge_energy = row["charge_energy_mwh"]
        discharge_energy = row["discharge_energy_mwh"]
        charge_time = row["charge_time_s"]
        discharge_time = row["discharge_time_s"]
        actual_charge_time = actual_row.get("charge_time_h")
        actual_discharge_time = actual_row.get("discharge_time_h")
        relaxed_clock_dialect = bool(
            record_clock_dialect
            and row.get("charge_time_is_clock")
            and row.get("discharge_time_is_clock")
        )
        capacity_tolerance = max(0.002, 0.001 * abs(float(charge_capacity))) if charge_capacity is not None else 0.0
        if relaxed_clock_dialect:
            capacity_tolerance = max(
                capacity_tolerance,
                _display_rounding_tolerance(float(charge_capacity), scale=1.0, minimum=0.0) * 2.5
                if charge_capacity is not None else 0.0,
            )
        compare(
            cycle_id,
            "charge capacity",
            actual_row.get("charge_capacity_mah"),
            charge_capacity,
            capacity_tolerance,
        )
        discharge_capacity_tolerance = max(0.002, 0.001 * abs(float(discharge_capacity))) if discharge_capacity is not None else 0.0
        if relaxed_clock_dialect:
            discharge_capacity_tolerance = max(
                discharge_capacity_tolerance,
                _display_rounding_tolerance(float(discharge_capacity), scale=1.0, minimum=0.0) * 2.5
                if discharge_capacity is not None else 0.0,
            )
        compare(
            cycle_id,
            "discharge capacity",
            actual_row.get("discharge_capacity_mah"),
            discharge_capacity,
            discharge_capacity_tolerance,
        )
        charge_energy_tolerance = max(0.01, 0.001 * abs(float(charge_energy))) if charge_energy is not None else 0.0
        if relaxed_clock_dialect:
            charge_energy_tolerance = max(
                charge_energy_tolerance,
                float(row.get("charge_energy_rounding_tolerance_mwh", 0.0)) * 2.5,
                abs(float(charge_energy)) * 0.005 if charge_energy is not None else 0.0,
            )
        compare(
            cycle_id,
            "charge energy",
            actual_row.get("charge_energy_mwh"),
            charge_energy,
            charge_energy_tolerance,
        )
        discharge_energy_tolerance = max(0.01, 0.001 * abs(float(discharge_energy))) if discharge_energy is not None else 0.0
        if relaxed_clock_dialect:
            discharge_energy_tolerance = max(
                discharge_energy_tolerance,
                float(row.get("discharge_energy_rounding_tolerance_mwh", 0.0)) * 2.5,
                abs(float(discharge_energy)) * 0.005 if discharge_energy is not None else 0.0,
            )
        compare(
            cycle_id,
            "discharge energy",
            actual_row.get("discharge_energy_mwh"),
            discharge_energy,
            discharge_energy_tolerance,
        )
        if charge_time is not None:
            compare(
                cycle_id,
                "charge time",
                float(actual_charge_time) * 3600.0,
                charge_time,
                time_tolerance_s,
            )
        if discharge_time is not None:
            compare(
                cycle_id,
                "discharge time",
                float(actual_discharge_time) * 3600.0,
                discharge_time,
                time_tolerance_s,
            )
        if (
            row["coulombic_efficiency_pct"] is not None
            and pd.notna(actual_row.get("coulombic_efficiency_pct"))
        ):
            efficiency_tolerance = 0.05
            if relaxed_clock_dialect:
                efficiency_tolerance = 0.5
            compare(
                cycle_id,
                "coulombic efficiency",
                actual_row.get("coulombic_efficiency_pct"),
                row["coulombic_efficiency_pct"],
                efficiency_tolerance,
            )
    state["cycle_summary_validated"] = True
