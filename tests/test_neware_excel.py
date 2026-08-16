from __future__ import annotations

import os
import shutil
import sys
import unittest
from datetime import datetime, timedelta
from pathlib import Path
from tempfile import TemporaryDirectory
from unittest import mock

import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

ROOT = Path(__file__).resolve().parents[1]
os.environ.setdefault("CELLXPLORER_DATA", str(ROOT / ".test-cellxplorer"))
sys.path.insert(0, str(ROOT / "backend"))

from app.config import CALC_VERSION
from app.db import Base
from app.models import Cell, SourceFile, Test, TestFile
from app.services import analysis_engine, cache, calc, chargeability, dcir, neware_excel, parsing, protocol, rate_capability


def load_tests(_loader, _tests, _pattern):
    """The concrete cases run from the coherent partition modules."""
    return unittest.TestSuite()


RECORD_HEADERS = [
    "DataPoint",
    "Cycle Index",
    "Step Index",
    "Step Type",
    "Time(min)",
    "Total Time(min)",
    "Current(mA)",
    "Voltage(V)",
    "Capacity(mAh)",
    "Spec. Cap.(mAh/g)",
    "Chg. Cap.(mAh)",
    "Chg. Spec. Cap.(mAh/g)",
    "DChg. Cap.(mAh)",
    "DChg. Spec. Cap.(mAh/g)",
    "Date",
    "Power(W)",
]

STEP_HEADERS = [
    "Cycle Index",
    "Step Index",
    "Step Number",
    "Step Type",
    "Step Time(min)",
    "Oneset Date",
    "End Date",
    "Capacity(mAh)",
    "Energy(Wh)",
    "Oneset Volt.(V)",
    "End Voltage(V)",
]


def _phase(step_type: str) -> str:
    if "DChg" in step_type:
        return "discharge"
    if "Chg" in step_type:
        return "charge"
    return "rest"


def _energy(power: list[float], total_time_min: list[float]) -> float:
    return sum(
        abs((power[index - 1] + power[index]) / 2.0)
        * max(0.0, total_time_min[index] - total_time_min[index - 1])
        / 60.0
        * 1000.0
        for index in range(1, len(power))
    )


def _write_synthetic_workbook(
    path: Path,
    *,
    include_step: bool = True,
    shuffled_records: bool = False,
    record_headers: list[str] | None = None,
) -> None:
    """Create a compact Neware-shaped workbook without private source data."""

    base = datetime(2026, 1, 1, 12, 0, 0)
    segments = [
        # Cycle 1 contains the same programmed step twice.  The second
        # occurrence must receive a distinct executed ``step`` id.
        (1, 1, "Rest", [0.0, 1.0], [3.50, 3.50], [0.0, 0.0], [0.0, 0.0]),
        (1, 2, "CC Chg", [0.0, 1.0, 2.0], [3.50, 3.60, 3.70], [0.0, 1.0, 2.0], [0.01] * 3),
        (1, 3, "Rest", [0.0, 1.0], [3.70, 3.70], [0.0, 0.0], [0.0, 0.0]),
        (1, 2, "CC Chg", [0.0, 1.0], [3.70, 3.80], [0.0, 0.5], [0.01] * 2),
        (1, 4, "CV Chg", [0.0, 1.0, 2.0], [3.80, 3.80, 3.80], [0.0, 0.2, 0.4], [0.005] * 3),
        (1, 5, "CC DChg", [0.0, 1.0, 2.0], [3.80, 3.20, 3.00], [0.0, 0.75, 1.5], [-0.008] * 3),
        (2, 1, "Rest", [0.0, 1.0], [3.00, 3.00], [0.0, 0.0], [0.0, 0.0]),
        (2, 2, "CC Chg", [0.0, 1.0, 2.0], [3.00, 3.40, 3.70], [0.0, 0.8, 1.6], [0.01] * 3),
        (2, 3, "CV Chg", [0.0, 1.0], [3.70, 3.70], [0.0, 0.2], [0.005] * 2),
        (2, 4, "CC DChg", [0.0, 1.0, 2.0], [3.70, 3.30, 3.00], [0.0, 0.6, 1.2], [-0.008] * 3),
    ]

    rows: list[dict[str, object]] = []
    summaries: list[list[object]] = []
    total_time_min = 0.0
    data_point = 1
    for cycle, step_index, step_type, times, voltages, capacities, power in segments:
        phase = _phase(step_type)
        current = 1.0 if phase == "charge" else -1.0 if phase == "discharge" else 0.0
        dates = [base + timedelta(minutes=total_time_min + value) for value in times]
        total_times = [total_time_min + value for value in times]
        for index, (time, total, voltage, capacity, watts, timestamp) in enumerate(
            zip(times, total_times, voltages, capacities, power, dates)
        ):
            rows.append(
                {
                    "DataPoint": data_point,
                    "Cycle Index": cycle,
                    "Step Index": step_index,
                    "Step Type": step_type,
                    "Time(min)": time,
                    "Total Time(min)": total,
                    "Current(mA)": current,
                    "Voltage(V)": voltage,
                    "Capacity(mAh)": capacity,
                    "Spec. Cap.(mAh/g)": capacity / 10.0,
                    "Chg. Cap.(mAh)": capacity if phase == "charge" else 0.0,
                    "Chg. Spec. Cap.(mAh/g)": capacity / 10.0 if phase == "charge" else 0.0,
                    "DChg. Cap.(mAh)": capacity if phase == "discharge" else 0.0,
                    "DChg. Spec. Cap.(mAh/g)": capacity / 10.0 if phase == "discharge" else 0.0,
                    "Date": timestamp,
                    "Power(W)": watts,
                }
            )
            data_point += 1

        summaries.append(
            [
                cycle,
                step_index,
                len(summaries) + 1,
                step_type,
                times[-1],
                dates[0],
                dates[-1],
                max(capacities) - min(capacities) if phase != "rest" else 0.0,
                _energy(power, total_times) / 1000.0,
                voltages[0],
                voltages[-1],
            ]
        )
        total_time_min = total_times[-1]

    workbook = Workbook()
    record_sheet = workbook.active
    record_sheet.title = "record"
    headers = record_headers or RECORD_HEADERS
    record_sheet.append(headers)
    source_rows = list(reversed(rows)) if shuffled_records else rows
    for row in source_rows:
        record_sheet.append([row.get(header.strip(), row.get(header)) for header in headers])

    if include_step:
        step_sheet = workbook.create_sheet("step")
        step_sheet.append(STEP_HEADERS)
        for summary in summaries:
            step_sheet.append(summary)

    workbook.save(path)


def _duration_text(seconds: float) -> str:
    milliseconds = int(round(float(seconds) * 1000.0))
    hours, remainder = divmod(milliseconds, 3_600_000)
    minutes, remainder = divmod(remainder, 60_000)
    whole_seconds, milliseconds = divmod(remainder, 1000)
    return f"{hours:02d}:{minutes:02d}:{whole_seconds:02d}.{milliseconds:03d}"


def _convert_to_duration_dialect(path: Path) -> None:
    """Convert the synthetic workbook to the observed ID/clock/kW dialect."""

    workbook = load_workbook(path)
    record = workbook["record"]
    record_headers = [cell.value for cell in record[1]]
    record_indices = {header: index + 1 for index, header in enumerate(record_headers)}
    for row in range(2, record.max_row + 1):
        time_cell = record.cell(row, record_indices["Time(min)"])
        total_time_cell = record.cell(row, record_indices["Total Time(min)"])
        power_cell = record.cell(row, record_indices["Power(W)"])
        time_cell.value = _duration_text(float(time_cell.value) * 60.0)
        total_time_cell.value = _duration_text(float(total_time_cell.value) * 60.0)
        power_cell.value = float(power_cell.value) / 1000.0
    for cell in record[1]:
        cell.value = {
            "Cycle Index": "Cycle ID",
            "Step Index": "Step ID",
            "Time(min)": "Time",
            "Total Time(min)": "Total Time",
            "Power(W)": "Power(kW)",
        }.get(cell.value, cell.value)

    step = workbook["step"]
    step_headers = [cell.value for cell in step[1]]
    step_indices = {header: index + 1 for index, header in enumerate(step_headers)}
    for row in range(2, step.max_row + 1):
        time_cell = step.cell(row, step_indices["Step Time(min)"])
        energy_cell = step.cell(row, step_indices["Energy(Wh)"])
        time_cell.value = _duration_text(float(time_cell.value) * 60.0)
        energy_cell.value = float(energy_cell.value) / 1000.0
    for cell in step[1]:
        cell.value = {
            "Cycle Index": "Cycle ID",
            "Step Index": "Step ID",
            "Step Time(min)": "Step Time",
            "Energy(Wh)": "Energy(kWh)",
        }.get(cell.value, cell.value)
    workbook.save(path)


def _add_duration_cycle_summary(path: Path, cycles: pd.DataFrame) -> None:
    workbook = load_workbook(path)
    cycle = workbook.create_sheet("cycle")
    cycle.append([
        "Cycle ID",
        "Chg. Cap.(mAh)",
        "DChg. Cap.(mAh)",
        "Chg. Time",
        "DChg. Time",
    ])
    for _, row in cycles.iterrows():
        charge_time = 0.0 if pd.isna(row["charge_time_h"]) else float(row["charge_time_h"]) * 3600.0
        discharge_time = 0.0 if pd.isna(row["discharge_time_h"]) else float(row["discharge_time_h"]) * 3600.0
        cycle.append([
            int(row["cycle"]),
            round(float(row["charge_capacity_mah"]), 3),
            round(float(row["discharge_capacity_mah"]), 3),
            _duration_text(charge_time),
            _duration_text(discharge_time),
        ])
    workbook.save(path)


def _convert_plan_to_duration_dialect(path: Path) -> None:
    workbook = load_workbook(path)
    test = workbook["test"]
    header_row = None
    time_column = None
    for row in test.iter_rows():
        for cell in row:
            if cell.value == "Step Time(min)":
                header_row = cell.row
                time_column = cell.column
                cell.value = "Step Time(hh:mm:ss.ms)"
                break
        if header_row is not None:
            break
    assert header_row is not None and time_column is not None
    for row in range(header_row + 1, test.max_row + 1):
        value = test.cell(row, time_column).value
        if isinstance(value, (int, float)) and not isinstance(value, bool):
            test.cell(row, time_column).value = _duration_text(float(value) * 60.0)
    for row in test.iter_rows():
        for cell in row:
            if cell.value == "Record settings":
                test.cell(cell.row, cell.column + 2).value = "60000ms/0.02V/0.1mA"
                break
    workbook.save(path)


TEST_PLAN_HEADERS = [
    "Step Index",
    "Step Name",
    "Step Time(min)",
    "Voltage(V)",
    "C-rate(C)",
    "Current(mA)",
    "Cut-off voltage (V)",
    "Cut-off C-rate(C)",
    "Cut-off curr.(mA)",
    "Energy(Wh)",
    "-ΔV(V)",
    "Power(W)",
    "Resistance(mΩ)",
    "Capacity(mAh)",
    "Record settings",
    "Aux.CH recording condition",
    "Max Vi(V)",
    "Min Vi(V)",
    "Max Ti(℃)",
    "Min Ti(℃)",
    "Segment record1",
    "Segment record2",
    "Current range (mA)",
]


def _write_metadata_workbook(path: Path, *, include_cycle: bool = False) -> None:
    _write_synthetic_workbook(path, include_step=False)
    workbook = load_workbook(path)
    test = workbook.create_sheet("test")
    test.append(["Test information"])
    test.append(["Start step ID", None, 1, "Volt. upper", None, "4.2V", "P/N", None, "PN-1"])
    test.append(["Cycle count", None, 3, "Volt. lower", None, "2.5V", "Builder", None, "Builder-1"])
    test.append(["Record settings", None, "5s/0.02V/0.1mA", None, None, None, "Remarks", None, "Remark-1"])
    test.append(["Voltage range", None, "4.2V", "Curr. lower", None, "-", "-", None, "-"])
    test.append(["Current range", None, "0-5mA", "Start time", None, datetime(2026, 1, 1, 12, 0, 0), "Barcode", None, "BAR-1"])
    test.append(["Active material", None, "10mg", "Nominal capacity", None, "10mAh", "", None, None])
    test.append([])
    test.append(["Step plan"])
    test.append(TEST_PLAN_HEADERS)
    test.append([1, "Rest", 2, None, None, None, None, None, None, None, None, None, None, None, "5s/0.02V/0.1mA"])
    test.append([2, "CC Chg", 10, None, 0.5, 5.0, 4.2, None, None, None, None, None, None, None, "5s/0.02V/0.1mA"])
    test.append([3, "CV Chg", 1, 4.2, 0.5, 5.0, None, None, 0.5, None, None, None, None, None, "5s/0.02V/0.1mA"])
    test.append([4, "CC DChg", 10, None, 0.5, 5.0, 2.5, None, None, None, None, None, None, None, "5s/0.02V/0.1mA"])
    test.append([5, "Cycle", "Start step ID:2", "Cycle count:3"])
    test.append([6, "CCCV DChg", None, None, 1.0, 10.0, 2.5, 0.05, 0.5, None, None, None, None, None, "5s/0.02V/0.1mA"])
    test.append([7, "CCCV Chg", 10, 4.2, 1.0, 10.0, None, 0.05, 0.5, None, None, None, None, None, "5s/0.02V/0.1mA"])
    test.append([8, "End"])
    unit = workbook.create_sheet("unit")
    unit.append(["synthetic.xlsx"])
    unit.append(["device", 1, 2, 3])
    unit.append(["Start time", None, datetime(2026, 1, 1, 12, 0, 0), None, "End time", None, datetime(2026, 1, 2, 12, 0, 0)])
    unit.append(["NDA file path", None, "C:/private/source.nda"])
    unit.append(["List of unit plans"])
    unit.append(["Time", "Current", "Voltage"])
    unit.append(["min", "mA", "V"])
    workbook.save(path)
    if include_cycle:
        raw = neware_excel.parse_timeseries(path)
        cycles = calc.per_cycle(raw)
        workbook = load_workbook(path)
        cycle = workbook.create_sheet("cycle")
        cycle.append([
            "Cycle Index",
            "Chg. Cap.(mAh)",
            "DChg. Cap.(mAh)",
            "Chg.-DChg. Eff(%)",
            "Chg. Energy(Wh)",
            "DChg. Energy(Wh)",
            "Chg. Time(min)",
            "DChg. Time(min)",
        ])
        for _, row in cycles.iterrows():
            cycle.append([
                int(row["cycle"]),
                round(float(row["charge_capacity_mah"]), 3),
                round(float(row["discharge_capacity_mah"]), 3),
                round(float(row["coulombic_efficiency_pct"]), 2),
                round(float(row["charge_energy_mwh"]) / 1000.0, 6),
                round(float(row["discharge_energy_mwh"]) / 1000.0, 6),
                round(float(row["charge_time_h"]) * 60.0, 3),
                round(float(row["discharge_time_h"]) * 60.0, 3),
            ])
        workbook.save(path)


def _add_declared_record_settings(path: Path, interval_s: float) -> None:
    workbook = load_workbook(path)
    test = workbook.create_sheet("test")
    test.append(["Record settings", None, f"{interval_s:g}s/0.01V/0mA"])
    test.append(["Step plan"])
    test.append(TEST_PLAN_HEADERS)
    workbook.save(path)


def _plan_row(
    step_index: int,
    step_name: str,
    *,
    step_time: float | str | None = None,
    voltage: float | None = None,
    rate: float | None = None,
    current: float | None = None,
    cutoff_voltage: float | None = None,
    cutoff_rate: float | None = None,
    cutoff_current: float | None = None,
) -> list[object]:
    row: list[object] = [None] * len(TEST_PLAN_HEADERS)
    values = {
        "Step Index": step_index,
        "Step Name": step_name,
        "Step Time(min)": step_time,
        "Voltage(V)": voltage,
        "C-rate(C)": rate,
        "Current(mA)": current,
        "Cut-off voltage (V)": cutoff_voltage,
        "Cut-off C-rate(C)": cutoff_rate,
        "Cut-off curr.(mA)": cutoff_current,
    }
    for header, value in values.items():
        row[TEST_PLAN_HEADERS.index(header)] = value
    return row


def _protocol_record(
    data_point: int,
    cycle: int,
    step_index: int,
    status: str,
    time_min: float,
    total_time_min: float,
    current_ma: float,
    voltage_v: float,
    capacity_mah: float,
    timestamp: datetime,
) -> dict[str, object]:
    charge = "Chg" in status and "DChg" not in status
    discharge = "DChg" in status
    return {
        "DataPoint": data_point,
        "Cycle Index": cycle,
        "Step Index": step_index,
        "Step Type": status,
        "Time(min)": time_min,
        "Total Time(min)": total_time_min,
        "Current(mA)": current_ma,
        "Voltage(V)": voltage_v,
        "Capacity(mAh)": capacity_mah,
        "Spec. Cap.(mAh/g)": capacity_mah / 0.1,
        "Chg. Cap.(mAh)": capacity_mah if charge else 0.0,
        "Chg. Spec. Cap.(mAh/g)": capacity_mah / 0.1 if charge else 0.0,
        "DChg. Cap.(mAh)": capacity_mah if discharge else 0.0,
        "DChg. Spec. Cap.(mAh/g)": capacity_mah / 0.1 if discharge else 0.0,
        "Date": timestamp,
        "Power(W)": voltage_v * current_ma / 1000.0,
    }


def _write_protocol_workbook(
    path: Path,
    *,
    plan_rows: list[list[object]],
    records: list[dict[str, object]],
    nominal_capacity_mah: float = 50.0,
    active_mass_mg: float = 100.0,
) -> None:
    """Write a deterministic Neware workbook for an analysis-family regression."""

    workbook = Workbook()
    record_sheet = workbook.active
    record_sheet.title = "record"
    record_sheet.append(RECORD_HEADERS)
    for record in records:
        record_sheet.append([record.get(header) for header in RECORD_HEADERS])

    test = workbook.create_sheet("test")
    test.append(["Test information"])
    test.append(["Start step ID", None, 1, "Volt. upper", None, "4.2V", "P/N", None, "ANALYSIS-039"])
    test.append(["Cycle count", None, len({int(row[0]) for row in plan_rows if row[0]}), "Volt. lower", None, "2.5V", "Builder", None, "CellXplorer"])
    test.append(["Record settings", None, "1s/0.01V/0.1mA", None, None, None, "Remarks", None, "Synthetic 039.4 fixture"])
    test.append(["Voltage range", None, "4.2V", "Curr. lower", None, "-", "-", None, "-"])
    test.append(["Current range", None, "0-50mA", "Start time", None, datetime(2026, 1, 1, 12, 0, 0), "Barcode", None, "SYNTH-0394"])
    test.append(["Active material", None, f"{active_mass_mg:g}mg", "Nominal capacity", None, f"{nominal_capacity_mah:g}mAh", "", None, None])
    test.append([])
    test.append(["Step plan"])
    test.append(TEST_PLAN_HEADERS)
    for row in plan_rows:
        test.append(row)

    unit = workbook.create_sheet("unit")
    unit.append([path.name])
    unit.append(["device", 1, 2, 3])
    unit.append(["Start time", None, datetime(2026, 1, 1, 12, 0, 0)])
    workbook.save(path)


class NewareExcelParserTests(unittest.TestCase):
    def test_unbounded_clock_duration_is_strict_and_unitless(self):
        self.assertAlmostEqual(
            neware_excel._clock_duration_seconds("1697:53:24.000"),
            1697 * 3600 + 53 * 60 + 24,
        )
        self.assertAlmostEqual(
            neware_excel._clock_duration_seconds("00:00:01.250"),
            1.25,
        )
        with self.assertRaises(ValueError):
            neware_excel._clock_duration_seconds("1:60:00")
        with self.assertRaises(ValueError):
            neware_excel._clock_duration_seconds("not-a-duration")

    def test_duration_id_and_kw_kwh_aliases_are_normalized_per_sheet(self):
        with TemporaryDirectory() as temporary:
            path = Path(temporary) / "duration-dialect.xlsx"
            _write_synthetic_workbook(path)
            _convert_to_duration_dialect(path)
            self.assertTrue(neware_excel.is_supported_workbook(path))
            frame = neware_excel.parse_timeseries(path)
            cycles = calc.per_cycle(frame)
            _add_duration_cycle_summary(path, cycles)
            neware_excel.validate_cycles(path, frame, cycles)
            with neware_excel._open(path) as workbook:
                summary = neware_excel._parse_cycle_summary(
                    neware_excel._sheet_by_name(workbook, "cycle", required=True)
                )

        np.testing.assert_allclose(frame["time_s"].iloc[:6], [0.0, 60.0, 0.0, 60.0, 120.0, 0.0])
        np.testing.assert_allclose(
            frame["total_time_s"].iloc[:6], [0.0, 60.0, 60.0, 120.0, 180.0, 180.0]
        )
        self.assertAlmostEqual(float(frame["power_w"].max()), 0.01)
        self.assertTrue(frame.attrs["neware_excel"]["step_summary_validated"])
        self.assertTrue(frame.attrs["neware_excel"]["step_summary_duration_validated"])
        self.assertTrue(frame.attrs["neware_excel"]["cycle_summary_validated"])
        self.assertIsNone(summary[0]["charge_energy_mwh"])
        self.assertIsNotNone(summary[0]["coulombic_efficiency_pct"])

    def test_duration_dialect_plan_and_millisecond_record_settings_are_read(self):
        with TemporaryDirectory() as temporary:
            path = Path(temporary) / "duration-metadata.xlsx"
            _write_metadata_workbook(path)
            _convert_plan_to_duration_dialect(path)
            metadata = neware_excel.read_metadata(path)

        self.assertEqual(
            metadata["Excel"]["Original"]["Test"]["RecordSettings"]["IntervalS"]["Value"],
            "60",
        )
        self.assertEqual(len(metadata["Step"]["Step_Info"]), 8)
        self.assertEqual(metadata["Step"]["Step_Info"]["Step1"]["Limit.Main.Time.Value"], "120000")
        self.assertIn(
            "StepTimeHhMmSsMs",
            metadata["Excel"]["Original"]["StepPlan"]["Step1"],
        )

    def test_numeric_dialect_cycle_tolerance_remains_locked(self):
        with TemporaryDirectory() as temporary:
            path = Path(temporary) / "strict-cycle.xlsx"
            _write_metadata_workbook(path, include_cycle=True)
            workbook = load_workbook(path)
            workbook["cycle"]["B2"] = float(workbook["cycle"]["B2"].value) + 0.1
            workbook.save(path)
            raw = neware_excel.parse_timeseries(path)
            with self.assertRaises(neware_excel.InvalidNewareExcelError):
                neware_excel.validate_cycles(path, raw, calc.per_cycle(raw))

    def test_unitless_numeric_duration_and_ambiguous_aliases_fail_closed(self):
        with TemporaryDirectory() as temporary:
            numeric_path = Path(temporary) / "numeric-duration.xlsx"
            _write_synthetic_workbook(numeric_path)
            _convert_to_duration_dialect(numeric_path)
            workbook = load_workbook(numeric_path)
            workbook["record"]["E2"] = 1.0
            workbook.save(numeric_path)
            with self.assertRaises(neware_excel.InvalidNewareExcelError):
                neware_excel.parse_timeseries(numeric_path)

            ambiguous_path = Path(temporary) / "ambiguous-alias.xlsx"
            _write_synthetic_workbook(ambiguous_path)
            workbook = load_workbook(ambiguous_path)
            workbook["record"].cell(1, workbook["record"].max_column + 1).value = "Cycle ID"
            workbook.save(ambiguous_path)
            with self.assertRaises(neware_excel.InvalidNewareExcelError):
                neware_excel.parse_timeseries(ambiguous_path)

            unsupported_path = Path(temporary) / "unsupported-power-unit.xlsx"
            _write_synthetic_workbook(unsupported_path)
            workbook = load_workbook(unsupported_path)
            workbook["record"]["P1"] = "Power(mW)"
            workbook.save(unsupported_path)
            with self.assertRaises(neware_excel.UnsupportedNewareExcelError):
                neware_excel.parse_timeseries(unsupported_path)

    def test_unitless_step_duration_is_reconciled(self):
        with TemporaryDirectory() as temporary:
            path = Path(temporary) / "duration-mismatch.xlsx"
            _write_synthetic_workbook(path)
            _convert_to_duration_dialect(path)
            workbook = load_workbook(path)
            workbook["step"]["E2"] = "999:00:00.000"
            workbook.save(path)
            with self.assertRaises(neware_excel.InvalidNewareExcelError):
                neware_excel.parse_timeseries(path)

    def test_numeric_step_duration_keeps_timestamp_span_validation(self):
        with TemporaryDirectory() as temporary:
            path = Path(temporary) / "numeric-timestamp-duration-mismatch.xlsx"
            _write_synthetic_workbook(path)
            workbook = load_workbook(path)
            first_record_end = workbook["record"]["O3"].value
            workbook["record"]["O3"] = first_record_end + timedelta(minutes=3)
            workbook["step"]["G2"] = workbook["record"]["O3"].value
            workbook.save(path)
            with self.assertRaises(neware_excel.InvalidNewareExcelError):
                neware_excel.parse_timeseries(path)

    def test_numeric_time_with_kwh_summary_keeps_strict_energy_tolerance(self):
        with TemporaryDirectory() as temporary:
            path = Path(temporary) / "numeric-time-kwh-energy.xlsx"
            _write_synthetic_workbook(path)
            workbook = load_workbook(path)
            step = workbook["step"]
            step["I1"] = "Energy(kWh)"
            for row in range(2, step.max_row + 1):
                value = step.cell(row, 9).value
                if value is not None:
                    step.cell(row, 9).value = 1.0
            step["I2"] = 1.000003
            workbook.save(path)
            with self.assertRaises(neware_excel.InvalidNewareExcelError):
                neware_excel.parse_timeseries(path)

    def test_valid_workbook_is_recognized(self):
        with TemporaryDirectory() as temporary:
            path = Path(temporary) / "synthetic.xlsx"
            _write_synthetic_workbook(path)
            self.assertTrue(neware_excel.is_supported_workbook(path))

    def test_unrelated_xlsx_is_rejected(self):
        with TemporaryDirectory() as temporary:
            path = Path(temporary) / "unrelated.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Sheet1"
            sheet.append(["Voltage", "Current"])
            sheet.append([3.7, 1.0])
            workbook.save(path)
            self.assertFalse(neware_excel.is_supported_workbook(path))
            with self.assertRaises(neware_excel.UnsupportedNewareExcelError):
                neware_excel.parse_timeseries(path)

    def test_metadata_rejects_unrelated_xlsx_before_labeling_source_format(self):
        with TemporaryDirectory() as temporary:
            path = Path(temporary) / "unrelated-metadata.xlsx"
            workbook = Workbook()
            workbook.save(path)

            with self.assertRaises(neware_excel.UnsupportedNewareExcelError):
                neware_excel.read_metadata(path)
            normalized = parsing.read_header_metadata(path)

        self.assertEqual(normalized["raw"], {})
        self.assertIn("error", normalized)
        self.assertNotIn("source_format", normalized)

    def test_missing_record_sheet_is_rejected(self):
        with TemporaryDirectory() as temporary:
            path = Path(temporary) / "missing-record.xlsx"
            workbook = Workbook()
            workbook.active.title = "unit"
            workbook.save(path)
            with self.assertRaises(neware_excel.UnsupportedNewareExcelError):
                neware_excel.parse_timeseries(path)

    def test_each_missing_required_record_header_is_rejected(self):
        for missing in neware_excel.REQUIRED_RECORD_HEADERS:
            with self.subTest(missing=missing), TemporaryDirectory() as temporary:
                path = Path(temporary) / "missing-column.xlsx"
                headers = [header for header in RECORD_HEADERS if header != missing]
                _write_synthetic_workbook(path, record_headers=headers)
                with self.assertRaises(neware_excel.UnsupportedNewareExcelError):
                    neware_excel.parse_timeseries(path)

    def test_duplicate_normalized_headers_are_rejected(self):
        with TemporaryDirectory() as temporary:
            path = Path(temporary) / "duplicate-header.xlsx"
            _write_synthetic_workbook(path, record_headers=RECORD_HEADERS + [" Voltage(V) "])
            with self.assertRaises(neware_excel.InvalidNewareExcelError):
                neware_excel.parse_timeseries(path)

    def test_corrupt_xlsx_is_rejected_with_domain_error(self):
        with TemporaryDirectory() as temporary:
            path = Path(temporary) / "corrupt.xlsx"
            path.write_bytes(b"not an xlsx zip")
            with self.assertRaises(neware_excel.InvalidNewareExcelError):
                neware_excel.parse_timeseries(path)

    def test_canonical_mapping_dtypes_and_auxiliary_columns(self):
        with TemporaryDirectory() as temporary:
            path = Path(temporary) / "synthetic.xlsx"
            _write_synthetic_workbook(path, shuffled_records=True)
            frame = neware_excel.parse_timeseries(path)

        self.assertEqual(len(frame), 25)
        self.assertEqual(frame["record_index"].tolist(), list(range(1, 26)))
        self.assertEqual(
            list(frame.columns),
            [
                "record_index",
                "cycle",
                "step",
                "step_index",
                "status",
                "time_s",
                "total_time_s",
                "voltage_v",
                "current_ma",
                "charge_capacity_mah",
                "discharge_capacity_mah",
                "charge_energy_mwh",
                "discharge_energy_mwh",
                "timestamp",
                "power_w",
                "capacity_mah",
                "specific_capacity_mah_g",
                "charge_specific_capacity_mah_g",
                "discharge_specific_capacity_mah_g",
            ],
        )
        for column in ("record_index", "cycle", "step", "step_index"):
            self.assertEqual(str(frame[column].dtype), "int64")
        for column in (
            "time_s",
            "total_time_s",
            "voltage_v",
            "current_ma",
            "charge_capacity_mah",
            "discharge_capacity_mah",
            "charge_energy_mwh",
            "discharge_energy_mwh",
            "power_w",
            "capacity_mah",
            "specific_capacity_mah_g",
            "charge_specific_capacity_mah_g",
            "discharge_specific_capacity_mah_g",
        ):
            self.assertEqual(str(frame[column].dtype), "float64")
        self.assertEqual(str(frame["timestamp"].dtype), "datetime64[ns]")
        self.assertTrue(pd.api.types.is_string_dtype(frame["status"]))
        self.assertEqual(
            set(frame["status"]),
            {"Rest", "CC_Chg", "CV_Chg", "CC_DChg"},
        )

    def test_time_units_and_duplicate_timestamps_are_preserved(self):
        with TemporaryDirectory() as temporary:
            path = Path(temporary) / "synthetic.xlsx"
            _write_synthetic_workbook(path)
            frame = neware_excel.parse_timeseries(path)

        np.testing.assert_allclose(frame["time_s"].iloc[:6], [0.0, 60.0, 0.0, 60.0, 120.0, 0.0])
        np.testing.assert_allclose(
            frame["total_time_s"].iloc[:6], [0.0, 60.0, 60.0, 120.0, 180.0, 180.0]
        )
        self.assertGreater(int(frame["timestamp"].duplicated().sum()), 0)
        self.assertTrue(frame["total_time_s"].is_monotonic_increasing)

    def test_invalid_required_values_are_rejected(self):
        cases = {
            "G2": "NaN",
            "H2": "not-a-number",
            "P2": "Infinity",
            "O2": "not-a-date",
        }
        for cell, value in cases.items():
            with self.subTest(cell=cell), TemporaryDirectory() as temporary:
                path = Path(temporary) / "invalid-value.xlsx"
                _write_synthetic_workbook(path, include_step=False)
                workbook = load_workbook(path)
                workbook["record"][cell] = value
                workbook.save(path)
                with self.assertRaises(neware_excel.InvalidNewareExcelError):
                    neware_excel.parse_timeseries(path)

    def test_verified_statuses_map_to_canonical_values(self):
        with TemporaryDirectory() as temporary:
            path = Path(temporary) / "statuses.xlsx"
            _write_synthetic_workbook(path, include_step=False)
            workbook = load_workbook(path)
            statuses = ["Rest", "CC Chg", "CCCV Chg", "CV Chg", "CC DChg", "CCCV DChg"]
            for row_number, status in enumerate(statuses, start=2):
                workbook["record"].cell(row_number, 4).value = status
            workbook.save(path)
            frame = neware_excel.parse_timeseries(path)

        self.assertTrue(
            {
                "Rest",
                "CC_Chg",
                "CCCV_Chg",
                "CV_Chg",
                "CC_DChg",
                "CCCV_DChg",
            }.issubset(set(frame["status"]))
        )

    def test_programmed_and_executed_steps_remain_distinct(self):
        with TemporaryDirectory() as temporary:
            path = Path(temporary) / "synthetic.xlsx"
            _write_synthetic_workbook(path)
            frame = neware_excel.parse_timeseries(path)

        repeated = frame.loc[
            (frame["cycle"] == 1) & (frame["step_index"] == 2), "step"
        ].unique()
        np.testing.assert_array_equal(repeated, [2, 4])
        self.assertEqual(frame["step"].nunique(), 10)
        self.assertEqual(frame.attrs["neware_excel"]["executed_step_count"], 10)
        self.assertTrue(frame.attrs["neware_excel"]["step_summary_validated"])
        self.assertTrue(frame["step"].is_monotonic_increasing)

    def test_time_reset_alone_starts_a_new_execution(self):
        with TemporaryDirectory() as temporary:
            path = Path(temporary) / "time-reset-only.xlsx"
            _write_synthetic_workbook(path, include_step=False)
            workbook = load_workbook(path)
            record_sheet = workbook["record"]
            # Rows 3 and 4 are consecutive source records. Make the second
            # one share every other boundary signal with the first while its
            # step-relative time resets from one minute to zero.
            record_sheet["C4"] = 1
            record_sheet["D4"] = "Rest"
            record_sheet["E4"] = 0.0
            record_sheet["G4"] = 0.0
            record_sheet["P4"] = 0.0
            workbook.save(path)
            frame = neware_excel.parse_timeseries(path)

        reset_rows = frame.loc[frame["record_index"].isin([2, 3])]
        self.assertEqual(reset_rows["cycle"].tolist(), [1, 1])
        self.assertEqual(reset_rows["step_index"].tolist(), [1, 1])
        self.assertEqual(reset_rows["status"].tolist(), ["Rest", "Rest"])
        np.testing.assert_allclose(reset_rows["time_s"], [60.0, 0.0])
        np.testing.assert_allclose(reset_rows["total_time_s"], [60.0, 60.0])
        self.assertEqual(reset_rows["step"].tolist(), [1, 2])
        self.assertTrue(frame["step"].is_monotonic_increasing)

    def test_energy_counters_reset_at_each_executed_step(self):
        with TemporaryDirectory() as temporary:
            path = Path(temporary) / "synthetic.xlsx"
            _write_synthetic_workbook(path)
            frame = neware_excel.parse_timeseries(path)

        charge_first = frame.groupby("step")["charge_energy_mwh"].first()
        discharge_first = frame.groupby("step")["discharge_energy_mwh"].first()
        np.testing.assert_allclose(charge_first.to_numpy(), 0.0)
        np.testing.assert_allclose(discharge_first.to_numpy(), 0.0)
        self.assertGreater(float(frame["charge_energy_mwh"].max()), 0.0)
        self.assertGreater(float(frame["discharge_energy_mwh"].max()), 0.0)
        charge_rows = frame["status"].str.contains("Chg") & ~frame["status"].str.contains("DChg")
        discharge_rows = frame["status"].str.contains("DChg")
        np.testing.assert_allclose(frame.loc[charge_rows, "discharge_energy_mwh"], 0.0)
        np.testing.assert_allclose(frame.loc[discharge_rows, "charge_energy_mwh"], 0.0)

    def test_calc_per_cycle_consumes_excel_frame_without_special_case(self):
        with TemporaryDirectory() as temporary:
            path = Path(temporary) / "synthetic.xlsx"
            _write_synthetic_workbook(path)
            frame = neware_excel.parse_timeseries(path)

        cycles = calc.per_cycle(frame)
        self.assertEqual(cycles["cycle"].tolist(), [1, 2])
        np.testing.assert_allclose(cycles["charge_capacity_mah"], [2.9, 1.8])
        np.testing.assert_allclose(cycles["discharge_capacity_mah"], [1.5, 1.2])
        np.testing.assert_allclose(cycles["charge_energy_mwh"], [2.0 / 3.0, 5.0 / 12.0])
        np.testing.assert_allclose(cycles["discharge_energy_mwh"], [4.0 / 15.0, 4.0 / 15.0])
        np.testing.assert_allclose(cycles["charge_time_h"], [5.0 / 60.0, 3.0 / 60.0])
        np.testing.assert_allclose(cycles["discharge_time_h"], [2.0 / 60.0, 2.0 / 60.0])
        np.testing.assert_allclose(cycles["cv_charge_capacity_mah"], [0.4, 0.2])
        np.testing.assert_allclose(cycles["cv_charge_time_h"], [2.0 / 60.0, 1.0 / 60.0])

    def test_missing_step_summary_keeps_raw_capability(self):
        with TemporaryDirectory() as temporary:
            path = Path(temporary) / "without-step.xlsx"
            _write_synthetic_workbook(path, include_step=False)
            self.assertTrue(neware_excel.is_supported_workbook(path))
            frame = neware_excel.parse_timeseries(path)

        self.assertEqual(len(frame), 25)
        self.assertFalse(frame.attrs["neware_excel"]["step_summary_available"])
        self.assertFalse(frame.attrs["neware_excel"]["step_summary_validated"])

    def test_step_summary_count_mismatch_is_rejected(self):
        with TemporaryDirectory() as temporary:
            path = Path(temporary) / "step-count-mismatch.xlsx"
            _write_synthetic_workbook(path)
            workbook = load_workbook(path)
            workbook["step"].delete_rows(workbook["step"].max_row)
            workbook.save(path)
            with self.assertRaises(neware_excel.InvalidNewareExcelError):
                neware_excel.parse_timeseries(path)

    def test_step_summary_identity_mismatch_is_rejected(self):
        with TemporaryDirectory() as temporary:
            path = Path(temporary) / "step-identity-mismatch.xlsx"
            _write_synthetic_workbook(path)
            workbook = load_workbook(path)
            workbook["step"]["A2"] = 99
            workbook.save(path)
            with self.assertRaises(neware_excel.InvalidNewareExcelError):
                neware_excel.parse_timeseries(path)

    def test_step_summary_duration_mismatch_is_rejected(self):
        with TemporaryDirectory() as temporary:
            path = Path(temporary) / "step-duration-mismatch.xlsx"
            _write_synthetic_workbook(path)
            workbook = load_workbook(path)
            workbook["step"]["E2"] = 100.0
            workbook.save(path)
            with self.assertRaises(neware_excel.InvalidNewareExcelError):
                neware_excel.parse_timeseries(path)

    def test_step_summary_rounding_uses_declared_record_cadence(self):
        with TemporaryDirectory() as temporary:
            path = Path(temporary) / "rounded-step-time.xlsx"
            _write_synthetic_workbook(path)
            _add_declared_record_settings(path, 60.0)
            workbook = load_workbook(path)
            # The raw segment lasts 60 seconds; a rounded 63-second summary is
            # accepted within the declared 60-second record cadence.
            workbook["step"]["E2"] = 1.05
            workbook.save(path)
            frame = neware_excel.parse_timeseries(path)

        self.assertTrue(frame.attrs["neware_excel"]["step_summary_validated"])

    def test_step_summary_does_not_infer_tolerance_from_sparse_timestamps(self):
        with TemporaryDirectory() as temporary:
            path = Path(temporary) / "sparse-step-time.xlsx"
            _write_synthetic_workbook(path)
            _add_declared_record_settings(path, 5.0)
            workbook = load_workbook(path)
            workbook["step"]["E2"] = 1.2
            workbook.save(path)
            with self.assertRaises(neware_excel.InvalidNewareExcelError):
                neware_excel.parse_timeseries(path)

    def test_unknown_status_is_rejected(self):
        with TemporaryDirectory() as temporary:
            path = Path(temporary) / "unknown-status.xlsx"
            _write_synthetic_workbook(path, include_step=False)
            workbook = load_workbook(path)
            workbook["record"]["D2"] = "Unknown"
            workbook.save(path)
            with self.assertRaises(neware_excel.InvalidNewareExcelError):
                neware_excel.parse_timeseries(path)

    def test_duplicate_datapoint_is_rejected(self):
        with TemporaryDirectory() as temporary:
            path = Path(temporary) / "duplicate-datapoint.xlsx"
            _write_synthetic_workbook(path, include_step=False)
            workbook = load_workbook(path)
            workbook["record"]["A3"] = 1
            workbook.save(path)
            with self.assertRaises(neware_excel.InvalidNewareExcelError):
                neware_excel.parse_timeseries(path)

    def test_total_time_decrease_is_rejected(self):
        with TemporaryDirectory() as temporary:
            path = Path(temporary) / "time-decrease.xlsx"
            _write_synthetic_workbook(path, include_step=False)
            workbook = load_workbook(path)
            workbook["record"]["F3"] = -1
            workbook.save(path)
            with self.assertRaises(neware_excel.InvalidNewareExcelError):
                neware_excel.parse_timeseries(path)

    def test_metadata_information_block_extracts_units_and_provenance(self):
        with TemporaryDirectory() as temporary:
            path = Path(temporary) / "metadata.xlsx"
            _write_metadata_workbook(path)
            metadata = neware_excel.read_metadata(path)

        head = metadata["Step"]["Head_Info"]
        self.assertEqual(head["Start_Step"]["Value"], "1")
        self.assertEqual(head["PN"]["Value"], "PN-1")
        self.assertEqual(head["Creator"]["Value"], "Builder-1")
        self.assertEqual(head["Remark"]["Value"], "Remark-1")
        self.assertEqual(head["SCQ"]["Value"], "10000.0")
        self.assertEqual(head["MultCap"]["Value"], "36000.0")
        self.assertEqual(metadata["Excel"]["Original"]["Test"]["StartTime"]["Value"], "2026-01-01 12:00:00")

    def test_missing_test_sheet_degrades_protocol_capability_without_blocking_raw_parse(self):
        with TemporaryDirectory() as temporary:
            path = Path(temporary) / "record-without-test.xlsx"
            _write_synthetic_workbook(path, include_step=False)
            workbook = load_workbook(path)
            unit = workbook.create_sheet("unit")
            unit.append(["synthetic.xlsx"])
            unit.append(["Start time", None, datetime(2026, 1, 1, 12, 0, 0)])
            workbook.save(path)

            raw = neware_excel.parse_timeseries(path)
            metadata = neware_excel.read_metadata(path)
            normalized = parsing.read_header_metadata(path)

        self.assertEqual(len(raw), 25)
        self.assertEqual(metadata["Step"]["Step_Info"], {})
        self.assertFalse(metadata["Excel"]["Capabilities"]["DeclaredProtocol"]["Value"])
        self.assertFalse(metadata["Excel"]["Capabilities"]["ProtocolConditions"]["Value"])
        self.assertEqual(normalized["source_format"], "Neware Excel")
        self.assertFalse(normalized["capabilities"]["DeclaredProtocol"])
        self.assertIsNone(normalized["nominal_capacity_mah"])
        self.assertEqual(normalized["start_time"], "2026-01-01 12:00:00")
        self.assertTrue(any("protocol condition expressions" in warning for warning in normalized["protocol_warnings"]))

    def test_metadata_protocol_plan_maps_explicit_fields_and_controls(self):
        with TemporaryDirectory() as temporary:
            path = Path(temporary) / "metadata.xlsx"
            _write_metadata_workbook(path)
            metadata = neware_excel.read_metadata(path)

        steps = metadata["Step"]["Step_Info"]
        expected_types = [4, 1, 3, 2, 5, 20, 7, 6]
        self.assertEqual([int(steps[f"Step{i}"]["Step_Type"]) for i in range(1, 9)], expected_types)
        for index, type_id in enumerate(expected_types, start=1):
            self.assertEqual(protocol.STEP_TYPES[type_id][0], protocol.STEP_TYPES[int(steps[f"Step{index}"]["Step_Type"])][0])

        cc = steps["Step2"]
        self.assertEqual(cc["Limit.Main.Curr.Value"], "5")
        self.assertEqual(cc["Limit.Main.Rate.Value"], "0.5")
        self.assertEqual(cc["Limit.Main.Stop_Volt.Value"], "42000")
        self.assertEqual(cc["Limit.Main.Time.Value"], "600000")
        self.assertEqual(cc["Record.Main.Time.Value"], "5000")
        self.assertEqual(cc["Record.Main.Volt.Value"], "200")
        self.assertEqual(cc["Protect.Main.Volt.Upper.Value"], "42000")
        self.assertEqual(cc["Protect.Main.Volt.Lower.Value"], "25000")
        self.assertEqual(steps["Step5"]["Limit.Other.Start_Step.Value"], "2")
        self.assertEqual(steps["Step5"]["Limit.Other.Cycle_Count.Value"], "3")
        self.assertEqual(steps["Step8"], {"Step_Type": "6"})

    def test_metadata_units_are_not_silently_reinterpreted(self):
        with TemporaryDirectory() as temporary:
            path = Path(temporary) / "wrong-unit.xlsx"
            _write_metadata_workbook(path)
            workbook = load_workbook(path)
            workbook["test"]["C7"] = "10V"
            workbook.save(path)
            with self.assertRaises(neware_excel.InvalidNewareExcelError):
                neware_excel.read_metadata(path)

    def test_metadata_value_groups_do_not_bleed_into_neighboring_labels(self):
        with TemporaryDirectory() as temporary:
            path = Path(temporary) / "missing-optional-values.xlsx"
            _write_metadata_workbook(path)
            workbook = load_workbook(path)
            workbook["test"]["C2"] = None
            workbook["test"]["F6"] = None
            workbook.remove(workbook["unit"])
            workbook.save(path)
            metadata = neware_excel.read_metadata(path)

        head = metadata["Step"]["Head_Info"]
        self.assertNotIn("Start_Step", head)
        self.assertEqual(head["Protect"]["Main"]["Volt"]["Upper"]["Value"], "42000")
        self.assertNotIn("StartTime", metadata["Excel"]["Original"]["Test"])

    def test_metadata_blank_group_does_not_consume_unsupported_neighbor_label(self):
        with TemporaryDirectory() as temporary:
            path = Path(temporary) / "blank-voltage-range.xlsx"
            _write_metadata_workbook(path)
            workbook = load_workbook(path)
            workbook["test"]["C5"] = None
            workbook.save(path)
            metadata = neware_excel.read_metadata(path)

        original_test = metadata["Excel"]["Original"]["Test"]
        self.assertNotIn("VoltageRange", original_test)

    def test_unit_optional_times_do_not_bleed_into_neighboring_labels(self):
        with TemporaryDirectory() as temporary:
            path = Path(temporary) / "missing-unit-start-time.xlsx"
            _write_metadata_workbook(path)
            workbook = load_workbook(path)
            workbook["unit"]["C3"] = None
            workbook.save(path)
            metadata = neware_excel.read_metadata(path)

        unit = metadata["Excel"]["Original"]["Unit"]
        self.assertNotIn("StartTime", unit)
        self.assertEqual(unit["EndTime"]["Value"], "2026-01-02 12:00:00")

    def test_numeric_record_date_is_rejected(self):
        with TemporaryDirectory() as temporary:
            path = Path(temporary) / "numeric-record-date.xlsx"
            _write_synthetic_workbook(path, include_step=False)
            workbook = load_workbook(path)
            workbook["record"]["O2"] = 123
            workbook["record"]["O2"].number_format = "General"
            workbook.save(path)
            with self.assertRaisesRegex(neware_excel.InvalidNewareExcelError, "Date"):
                neware_excel.parse_timeseries(path)

    def test_numeric_record_date_string_is_rejected(self):
        with TemporaryDirectory() as temporary:
            path = Path(temporary) / "numeric-record-date-string.xlsx"
            _write_synthetic_workbook(path, include_step=False)
            workbook = load_workbook(path)
            workbook["record"]["O2"] = "20260101"
            workbook.save(path)
            with self.assertRaisesRegex(neware_excel.InvalidNewareExcelError, "Date"):
                neware_excel.parse_timeseries(path)

    def test_fast_and_reference_paths_match_exactly(self):
        if neware_excel._fastexcel is None:
            self.skipTest("fastexcel is not installed")
        with TemporaryDirectory() as temporary:
            path = Path(temporary) / "parity.xlsx"
            _write_synthetic_workbook(path)
            fast = neware_excel.parse_timeseries(path)
            fast_reader = neware_excel._fastexcel
            calamine_reader = neware_excel._python_calamine
            neware_excel._fastexcel = None
            neware_excel._python_calamine = None
            try:
                reference = neware_excel.parse_timeseries(path)
            finally:
                neware_excel._fastexcel = fast_reader
                neware_excel._python_calamine = calamine_reader

        self.assertTrue(fast.equals(reference))
        self.assertEqual(fast.attrs["neware_excel"], reference.attrs["neware_excel"])

    def test_calamine_and_reference_paths_match_exactly(self):
        if neware_excel._python_calamine is None:
            self.skipTest("python-calamine is not installed")
        with TemporaryDirectory() as temporary:
            path = Path(temporary) / "calamine-parity.xlsx"
            _write_synthetic_workbook(path)
            fast_reader = neware_excel._fastexcel
            calamine_reader = neware_excel._python_calamine
            neware_excel._fastexcel = None
            try:
                calamine = neware_excel.parse_timeseries(path)
                neware_excel._python_calamine = None
                reference = neware_excel.parse_timeseries(path)
            finally:
                neware_excel._fastexcel = fast_reader
                neware_excel._python_calamine = calamine_reader

        self.assertTrue(calamine.equals(reference))
        self.assertEqual(calamine.attrs["neware_excel"], reference.attrs["neware_excel"])

    def test_calamine_falls_back_when_unavailable(self):
        with TemporaryDirectory() as temporary:
            path = Path(temporary) / "calamine-fallback.xlsx"
            _write_synthetic_workbook(path)
            original_fast = neware_excel._fastexcel
            original_calamine = neware_excel._python_calamine
            neware_excel._fastexcel = None
            neware_excel._python_calamine = None
            try:
                frame = neware_excel.parse_timeseries(path)
            finally:
                neware_excel._fastexcel = original_fast
                neware_excel._python_calamine = original_calamine

        self.assertEqual(len(frame), 25)
        self.assertTrue(frame.attrs["neware_excel"]["step_summary_validated"])

    def test_calamine_native_duration_representation_falls_back_to_openpyxl(self):
        if neware_excel._python_calamine is None:
            self.skipTest("python-calamine is not installed")
        with TemporaryDirectory() as temporary:
            path = Path(temporary) / "calamine-native-duration.xlsx"
            _write_synthetic_workbook(path, include_step=False)
            workbook = load_workbook(path)
            record = workbook["record"]
            indices = {cell.value: cell.column for cell in record[1]}
            for row in range(2, record.max_row + 1):
                time_cell = record.cell(row, indices["Time(min)"])
                total_time_cell = record.cell(row, indices["Total Time(min)"])
                power_cell = record.cell(row, indices["Power(W)"])
                time_cell.value = timedelta(minutes=float(time_cell.value))
                total_time_cell.value = timedelta(minutes=float(total_time_cell.value))
                power_cell.value = float(power_cell.value) / 1000.0
            record.cell(1, indices["Time(min)"]).value = "Time"
            record.cell(1, indices["Total Time(min)"]).value = "Total Time"
            record.cell(1, indices["Power(W)"]).value = "Power(kW)"
            workbook.save(path)

            original_fast = neware_excel._fastexcel
            neware_excel._fastexcel = None
            try:
                self.assertIsNone(neware_excel._parse_calamine_timeseries(path))
                with mock.patch.object(
                    neware_excel,
                    "_parse_records",
                    wraps=neware_excel._parse_records,
                ) as reference_parse:
                    frame = neware_excel.parse_timeseries(path)
            finally:
                neware_excel._fastexcel = original_fast

        self.assertTrue(reference_parse.called)
        np.testing.assert_allclose(frame["time_s"].iloc[:6], [0.0, 60.0, 0.0, 60.0, 120.0, 0.0])

    def test_calamine_open_failure_falls_back_to_openpyxl(self):
        if neware_excel._python_calamine is None:
            self.skipTest("python-calamine is not installed")
        with TemporaryDirectory() as temporary:
            path = Path(temporary) / "calamine-open-failure.xlsx"
            _write_synthetic_workbook(path, include_step=False)
            original_fast = neware_excel._fastexcel
            neware_excel._fastexcel = None
            try:
                with mock.patch.object(
                    neware_excel.pd,
                    "ExcelFile",
                    side_effect=ValueError("simulated calamine failure"),
                ), mock.patch.object(
                    neware_excel,
                    "_parse_records",
                    wraps=neware_excel._parse_records,
                ) as reference_parse:
                    frame = neware_excel.parse_timeseries(path)
            finally:
                neware_excel._fastexcel = original_fast

        self.assertTrue(reference_parse.called)
        self.assertEqual(len(frame), 25)

    def test_fast_path_rejects_ambiguous_normalized_record_sheets(self):
        if neware_excel._fastexcel is None:
            self.skipTest("fastexcel is not installed")
        with TemporaryDirectory() as temporary:
            path = Path(temporary) / "ambiguous-record-sheets.xlsx"
            _write_synthetic_workbook(path, include_step=False)
            workbook = load_workbook(path)
            duplicate = workbook.copy_worksheet(workbook["record"])
            duplicate.title = " record "
            workbook.save(path)
            with self.assertRaisesRegex(neware_excel.InvalidNewareExcelError, "ambiguous"):
                neware_excel.parse_timeseries(path)

    def test_fast_path_does_not_ignore_rows_populated_outside_projection(self):
        if neware_excel._fastexcel is None:
            self.skipTest("fastexcel is not installed")
        with TemporaryDirectory() as temporary:
            path = Path(temporary) / "malformed-extra-column.xlsx"
            _write_synthetic_workbook(path, include_step=False)
            workbook = load_workbook(path)
            workbook["record"]["Q27"] = "unexpected"
            workbook.save(path)
            with self.assertRaises(neware_excel.InvalidNewareExcelError):
                neware_excel.parse_timeseries(path)

    def test_fast_path_rejects_integer_identifier_overflow(self):
        if neware_excel._fastexcel is None:
            self.skipTest("fastexcel is not installed")
        for column in ("A", "B", "C"):
            for value in (1e20, 2**63 - 1, 2**63):
                with self.subTest(column=column, value=value), TemporaryDirectory() as temporary:
                    path = Path(temporary) / "identifier-overflow.xlsx"
                    _write_synthetic_workbook(path, include_step=False)
                    workbook = load_workbook(path)
                    workbook["record"][f"{column}2"] = value
                    workbook.save(path)

                    fast_error = None
                    try:
                        neware_excel.parse_timeseries(path)
                    except neware_excel.InvalidNewareExcelError as exc:
                        fast_error = str(exc)
                    self.assertIsNotNone(fast_error)

                    fast_reader = neware_excel._fastexcel
                    calamine_reader = neware_excel._python_calamine
                    neware_excel._fastexcel = None
                    neware_excel._python_calamine = None
                    try:
                        with self.assertRaises(neware_excel.InvalidNewareExcelError) as reference:
                            neware_excel.parse_timeseries(path)
                    finally:
                        neware_excel._fastexcel = fast_reader
                        neware_excel._python_calamine = calamine_reader
                    self.assertEqual(fast_error, str(reference.exception))

    def test_numeric_metadata_start_time_is_rejected(self):
        with TemporaryDirectory() as temporary:
            path = Path(temporary) / "numeric-metadata-start-time.xlsx"
            _write_metadata_workbook(path)
            workbook = load_workbook(path)
            workbook["test"]["F6"] = 123
            workbook["test"]["F6"].number_format = "General"
            workbook.save(path)
            with self.assertRaisesRegex(neware_excel.InvalidNewareExcelError, "Start time"):
                neware_excel.read_metadata(path)

    def test_shared_metadata_normalization_and_capabilities(self):
        with TemporaryDirectory() as temporary:
            path = Path(temporary) / "metadata.xlsx"
            _write_metadata_workbook(path)
            result = parsing.read_header_metadata(path)

        self.assertEqual(result["source_format"], "Neware Excel")
        self.assertEqual(result["start_time"], "2026-01-01 12:00:00")
        self.assertEqual(result["active_mass_mg"], 10.0)
        self.assertEqual(result["nominal_capacity_mah"], 10.0)
        self.assertEqual(result["protection_voltage_upper_v"], 4.2)
        self.assertEqual(result["protection_voltage_lower_v"], 2.5)
        self.assertEqual(result["record_interval_s"], 5.0)
        self.assertFalse(result["capabilities"]["ProtocolConditions"])

    def test_protocol_reconstruction_and_signature_use_excel_flattened_plan(self):
        with TemporaryDirectory() as temporary:
            first = Path(temporary) / "first.xlsx"
            second = Path(temporary) / "second.xlsx"
            _write_metadata_workbook(first)
            _write_metadata_workbook(second)
            first_result = parsing.read_header_metadata(first)
            second_result = parsing.read_header_metadata(second)
            first_protocol = protocol.reconstruct_protocol(first_result["raw"], 10.0)
            second_protocol = protocol.reconstruct_protocol(second_result["raw"], 10.0)

            workbook = load_workbook(second)
            workbook["test"]["G13"] = 4.1
            workbook.save(second)
            changed = parsing.read_header_metadata(second)
            changed_protocol = protocol.reconstruct_protocol(changed["raw"], 10.0)

        self.assertEqual(first_protocol["n_steps"], 8)
        self.assertEqual(first_protocol["n_executable_steps"], 6)
        self.assertEqual(first_protocol["signature"], second_protocol["signature"])
        self.assertNotEqual(first_protocol["signature"], changed_protocol["signature"])
        self.assertEqual(first_protocol["summary"]["protection_windows"], [{"lower_v": 2.5, "upper_v": 4.2}])
        self.assertEqual(first_protocol["steps"][4]["conditions"], [])
        self.assertTrue(any("protocol condition expressions" in warning for warning in first_protocol["warnings"]))
        self.assertEqual(chargeability.detect_candidates(first_protocol), [])

    def test_three_rate_excel_plan_reaches_rate_capability_pairing_seam(self):
        with TemporaryDirectory() as temporary:
            path = Path(temporary) / "three-rate-metadata.xlsx"
            _write_metadata_workbook(path)
            workbook = load_workbook(path)
            test = workbook["test"]
            test.delete_rows(11, test.max_row - 10)

            def plan_row(
                step_index: int,
                name: str,
                *,
                step_time: object = None,
                voltage: object = None,
                rate: object = None,
                current: object = None,
                cutoff_voltage: object = None,
                cutoff_rate: object = None,
                cutoff_current: object = None,
            ) -> list[object]:
                row: list[object] = [None] * len(TEST_PLAN_HEADERS)
                row[0] = step_index
                row[1] = name
                row[2] = step_time
                row[3] = voltage
                row[4] = rate
                row[5] = current
                row[6] = cutoff_voltage
                row[7] = cutoff_rate
                row[8] = cutoff_current
                return row

            rows: list[list[object]] = []
            step_index = 1
            for charge_rate in (0.2, 0.5, 1.0):
                rows.append(plan_row(step_index, "Rest", step_time=1))
                step_index += 1
                rows.append(
                    plan_row(
                        step_index,
                        "CC Chg",
                        rate=charge_rate,
                        current=charge_rate * 10.0,
                        voltage=4.2,
                    )
                )
                step_index += 1
                rows.append(
                    plan_row(
                        step_index,
                        "CV Chg",
                        rate=charge_rate,
                        current=charge_rate * 10.0,
                        voltage=4.2,
                        cutoff_current=0.5,
                    )
                )
                step_index += 1
                rows.append(
                    plan_row(
                        step_index,
                        "CC DChg",
                        rate=1.0,
                        current=10.0,
                        cutoff_voltage=2.5,
                    )
                )
                step_index += 1
            rows.append(plan_row(step_index, "End"))
            for row in rows:
                test.append(row)
            workbook.save(path)

            metadata = parsing.read_header_metadata(path)
            reconstructed = protocol.reconstruct_protocol(
                metadata["raw"], metadata["nominal_capacity_mah"]
            )
            pairs = rate_capability.build_rate_pairs(reconstructed)

        self.assertEqual(len(pairs), 3)
        self.assertEqual(
            [pair["charge"]["measurement_step_index"] for pair in pairs],
            [2, 6, 10],
        )
        self.assertEqual(
            [round(pair["charge_rate_c"], 3) for pair in pairs],
            [0.2, 0.5, 1.0],
        )
        self.assertEqual(
            [pair["charge"]["direction"] for pair in pairs],
            ["charge", "charge", "charge"],
        )
        self.assertEqual(
            [pair["discharge"]["direction"] for pair in pairs],
            ["discharge", "discharge", "discharge"],
        )
        self.assertEqual(
            [pair["charge"]["step_indices"] for pair in pairs],
            [[2, 3], [6, 7], [10, 11]],
        )
        self.assertEqual([pair["upper_voltage_v"] for pair in pairs], [4.2, 4.2, 4.2])
        self.assertEqual([pair["lower_voltage_v"] for pair in pairs], [2.5, 2.5, 2.5])

    def test_metadata_read_does_not_parse_large_record_sheet(self):
        with TemporaryDirectory() as temporary:
            path = Path(temporary) / "metadata.xlsx"
            _write_metadata_workbook(path)
            with mock.patch.object(neware_excel, "_parse_records", side_effect=AssertionError("record scan")):
                metadata = neware_excel.read_metadata(path)
        self.assertTrue(metadata["Excel"]["Capabilities"]["DeclaredProtocol"]["Value"])

    def test_parser_dispatch_preserves_binary_and_excel_boundaries(self):
        binary_frame = pd.DataFrame({"Index": [1], "Cycle": [1], "Status": ["Rest"]})
        with mock.patch.object(parsing.NewareNDA, "read", return_value=binary_frame) as binary_read:
            result = parsing.parse_timeseries("source.nda")
        binary_read.assert_called_once()
        self.assertEqual(result.loc[0, "record_index"], 1)

        with mock.patch.object(neware_excel, "parse_timeseries", return_value=pd.DataFrame({"cycle": [1]})) as excel_read:
            result = parsing.parse_timeseries("source.xlsx")
        excel_read.assert_called_once()
        self.assertEqual(result.loc[0, "cycle"], 1)
        with self.assertRaises(parsing.UnsupportedSourceFormatError):
            parsing.parse_timeseries("source.csv")

    def test_parser_bundle_version_is_deterministic_and_persistable(self):
        self.assertEqual(neware_excel.EXCEL_PARSER_REVISION, 6)
        self.assertIn(parsing.NEWARE_NDA_VERSION, parsing.PARSER_VERSION)
        self.assertIn(f"cxp{neware_excel.EXCEL_PARSER_REVISION}", parsing.PARSER_VERSION)
        self.assertLessEqual(len(parsing.PARSER_VERSION), 30)
        self.assertEqual(parsing.PARSER_VERSION, f"{parsing.NEWARE_NDA_VERSION}-cxp{neware_excel.EXCEL_PARSER_REVISION}")
        self.assertNotEqual(
            parsing.PARSER_VERSION,
            f"{parsing.NEWARE_NDA_VERSION}-cxp{neware_excel.EXCEL_PARSER_REVISION + 1}",
        )
        self.assertEqual(CALC_VERSION, "1.6.1")

    def test_cycle_summary_validation_accepts_rounded_values(self):
        with TemporaryDirectory() as temporary:
            path = Path(temporary) / "cycle-summary.xlsx"
            _write_metadata_workbook(path, include_cycle=True)
            raw = neware_excel.parse_timeseries(path)
            cycles = calc.per_cycle(raw)
            neware_excel.validate_cycles(path, raw, cycles)
        self.assertTrue(raw.attrs["neware_excel"]["cycle_summary_validated"])

    def test_cycle_summary_rejects_non_finite_calculated_values_even_at_zero(self):
        with TemporaryDirectory() as temporary:
            path = Path(temporary) / "cycle-non-finite.xlsx"
            _write_metadata_workbook(path, include_cycle=True)
            workbook = load_workbook(path)
            workbook["cycle"]["B2"] = 0.0
            workbook.save(path)
            raw = neware_excel.parse_timeseries(path)
            cycles = calc.per_cycle(raw)
            cycles.loc[0, "charge_capacity_mah"] = np.inf
            with self.assertRaises(neware_excel.InvalidNewareExcelError):
                neware_excel.validate_cycles(path, raw, cycles)

    def test_cycle_summary_rejects_nan_calculated_capacity_even_at_zero(self):
        with TemporaryDirectory() as temporary:
            path = Path(temporary) / "cycle-nan.xlsx"
            _write_metadata_workbook(path, include_cycle=True)
            workbook = load_workbook(path)
            workbook["cycle"]["B2"] = 0.0
            workbook.save(path)
            raw = neware_excel.parse_timeseries(path)
            cycles = calc.per_cycle(raw)
            cycles.loc[0, "charge_capacity_mah"] = np.nan
            with self.assertRaises(neware_excel.InvalidNewareExcelError):
                neware_excel.validate_cycles(path, raw, cycles)

    def test_cycle_summary_identity_capacity_energy_and_time_mismatches_fail(self):
        mutations = {
            "identity": ("A2", 99),
            "capacity": ("B2", 999.0),
            "energy": ("E2", 999.0),
            "time": ("G2", 999.0),
        }
        for name, (cell, value) in mutations.items():
            with self.subTest(name=name), TemporaryDirectory() as temporary:
                path = Path(temporary) / f"cycle-{name}.xlsx"
                _write_metadata_workbook(path, include_cycle=True)
                workbook = load_workbook(path)
                workbook["cycle"][cell] = value
                workbook.save(path)
                raw = neware_excel.parse_timeseries(path)
                cycles = calc.per_cycle(raw)
                with self.assertRaises(neware_excel.InvalidNewareExcelError):
                    neware_excel.validate_cycles(path, raw, cycles)

    def test_missing_cycle_summary_is_explicitly_non_validating(self):
        with TemporaryDirectory() as temporary:
            path = Path(temporary) / "without-cycle.xlsx"
            _write_metadata_workbook(path)
            raw = neware_excel.parse_timeseries(path)
            cycles = calc.per_cycle(raw)
            neware_excel.validate_cycles(path, raw, cycles)
        self.assertFalse(raw.attrs["neware_excel"]["cycle_summary_available"])
        self.assertFalse(raw.attrs["neware_excel"]["cycle_summary_validated"])

    def test_cache_build_and_write_behind_validate_excel_before_publication(self):
        with TemporaryDirectory() as temporary:
            path = Path(temporary) / "cached.xlsx"
            _write_metadata_workbook(path, include_cycle=True)
            file_hash = parsing.compute_hash(path)
            identity = parsing.parser_identity(path)
            cache_directory = cache.raw_path(file_hash, identity).parent
            try:
                with mock.patch.object(parsing, "validate_parsed_output", wraps=parsing.validate_parsed_output) as validate:
                    cache.build(file_hash, path)
                validate.assert_called_once()
                self.assertIsNotNone(cache.load_raw(file_hash, identity))
                self.assertIsNotNone(cache.load_cycles(file_hash, identity, cache.CALC_VERSION))
            finally:
                if cache_directory.exists():
                    import shutil
                    shutil.rmtree(cache_directory)

        with TemporaryDirectory() as temporary:
            path = Path(temporary) / "write-behind.xlsx"
            _write_metadata_workbook(path, include_cycle=True)
            file_hash = parsing.compute_hash(path)
            identity = parsing.parser_identity(path)
            cache_directory = cache.raw_path(file_hash, identity).parent
            try:
                with mock.patch.object(parsing, "validate_parsed_output", wraps=parsing.validate_parsed_output) as validate:
                    cache.build_write_behind(file_hash, path)
                validate.assert_called_once()
                cache.wait_for_pending(file_hash)
                self.assertTrue(cache.raw_path(file_hash, identity).exists())
                self.assertTrue(cache.cycles_path(file_hash, identity).exists())
            finally:
                if cache_directory.exists():
                    import shutil
                    shutil.rmtree(cache_directory)

    def test_cycle_cache_derivation_from_existing_raw_does_not_reopen_excel(self):
        with TemporaryDirectory() as temporary:
            path = Path(temporary) / "derive.xlsx"
            _write_metadata_workbook(path, include_cycle=True)
            file_hash = parsing.compute_hash(path)
            identity = parsing.parser_identity(path)
            cache_directory = cache.raw_path(file_hash, identity).parent
            try:
                cache.build(file_hash, path)
                cache.cycles_path(file_hash, identity).unlink()
                with mock.patch.object(parsing, "validate_parsed_output") as validate:
                    cycles = cache.load_cycles(file_hash, identity, cache.CALC_VERSION)
                self.assertIsNotNone(cycles)
                validate.assert_not_called()
            finally:
                if cache_directory.exists():
                    import shutil
                    shutil.rmtree(cache_directory)


class NewareExcelAnalysisIntegrationTests(unittest.TestCase):
    """Exercise registered Excel sources through the existing analysis services."""

    def setUp(self):
        engine = create_engine(
            "sqlite://",
            connect_args={"check_same_thread": False},
            poolclass=StaticPool,
        )
        Base.metadata.create_all(engine)
        self.db = sessionmaker(bind=engine, autoflush=False, expire_on_commit=False)()
        self.cache_hashes: list[str] = []

    def tearDown(self):
        self.db.close()
        for file_hash in self.cache_hashes:
            shutil.rmtree(cache.raw_path(file_hash).parent, ignore_errors=True)

    def _register(self, path: Path) -> tuple[Cell, SourceFile, dict, pd.DataFrame]:
        metadata = parsing.read_header_metadata(path)
        raw = neware_excel.parse_timeseries(path)
        file_hash = parsing.compute_hash(path)
        build_info = cache.build(file_hash, path)
        self.cache_hashes.append(file_hash)

        source = SourceFile(
            hash=file_hash,
            path=str(path),
            filename=path.name,
            size=path.stat().st_size,
            ext=path.suffix.casefold().lstrip("."),
            parse_status="parsed",
            parser_version=build_info["parser_version"],
            row_count=len(raw),
            cycle_count=int(raw["cycle"].nunique()),
            header_meta=metadata["raw"],
            start_time=metadata.get("start_time"),
            active_mass_mg=metadata.get("active_mass_mg"),
            nominal_capacity_mah=metadata.get("nominal_capacity_mah"),
            capacity_summary_status="ready",
        )
        cell = Cell(name=f"Excel integration {len(self.cache_hashes)}")
        test = Test(cell=cell, name="internal source chain")
        self.db.add_all([cell, source, test])
        self.db.flush()
        self.db.add(TestFile(test_id=test.id, file_id=source.id, position=0))
        self.db.commit()
        return cell, source, metadata, raw

    @staticmethod
    def _spec(cell: Cell) -> dict:
        spec = analysis_engine.default_spec("Excel integration")
        spec["selection"]["entries"] = [{"kind": "cell", "ref_id": cell.id}]
        return spec

    def test_registered_excel_feeds_cycles_time_capacity_and_repeated_steps(self):
        with TemporaryDirectory() as temporary:
            path = Path(temporary) / "analysis.xlsx"
            _write_metadata_workbook(path)
            cell, source, metadata, raw = self._register(path)
            spec = self._spec(cell)

            cycles = analysis_engine.compute(self.db, spec, None)
            time_capacity = analysis_engine.compute_time_capacity(self.db, spec, None)

            signature = protocol.reconstruct_protocol(
                metadata["raw"], source.nominal_capacity_mah
            )["signature"]
            spec["protocol_segments"] = [
                {
                    "id": "repeated-charge",
                    "name": "Repeated charge step",
                    "targets": [
                        {"protocol_signature": signature, "step_indices": [2]}
                    ],
                }
            ]
            spec["computation"]["steps"] = {
                "series": [
                    {
                        "id": "repeated-charge-series",
                        "cell_id": cell.id,
                        "segment_id": "repeated-charge",
                    }
                ],
                "mode": "union",
            }
            steps = analysis_engine.compute_steps(self.db, spec, None)

        cycle_series = cycles["cell_series"][0]
        self.assertEqual(source.ext, "xlsx")
        self.assertEqual(cycle_series["x"], [1, 2])
        np.testing.assert_allclose(
            cycle_series["quantities"]["charge_capacity_mah"], [2.9, 1.8]
        )
        self.assertEqual(cycles["sources"][0]["file_hashes"], [source.hash])

        trace = time_capacity["cell_traces"][0]
        self.assertEqual(len(trace["time_s"]), len(raw))
        self.assertEqual(set(trace["source_filename"]), {path.name})
        times = [value for value in trace["time_s"] if value is not None]
        self.assertTrue(all(second >= first for first, second in zip(times, times[1:])))

        step_series = steps["cell_series"][0]
        self.assertEqual(step_series["n_blocks"], 3)
        self.assertEqual(step_series["x_cycle"][:2], [1, 1])
        self.assertEqual(
            [item["step_start"] for item in step_series["block_meta"][:2]], [2, 2]
        )

    def test_registered_excel_dcir_fixture_uses_current_detector_and_occurrences(self):
        plan = [
            _plan_row(1, "Rest", step_time=30),
            _plan_row(
                2,
                "CC DChg",
                step_time=0.5,
                rate=1.0,
                current=-50.0,
                cutoff_voltage=3.4,
            ),
            _plan_row(3, "End"),
        ]
        records: list[dict[str, object]] = []
        base = datetime(2026, 1, 1, 12, 0, 0)
        total_min = 0.0
        data_point = 1
        for cycle in range(1, 4):
            for step_index, status, current, duration, voltages in (
                (1, "Rest", 0.0, 30.0, (3.50, 3.50)),
                (2, "CC DChg", -50.0, 0.5, (3.50, 3.40)),
            ):
                records.append(
                    _protocol_record(
                        data_point,
                        cycle,
                        step_index,
                        status,
                        0.0,
                        total_min,
                        current,
                        voltages[0],
                        0.0,
                        base + timedelta(minutes=total_min),
                    )
                )
                data_point += 1
                records.append(
                    _protocol_record(
                        data_point,
                        cycle,
                        step_index,
                        status,
                        duration,
                        total_min + duration,
                        current,
                        voltages[1],
                        0.0,
                        base + timedelta(minutes=total_min + duration),
                    )
                )
                data_point += 1
                total_min += duration

        with TemporaryDirectory() as temporary:
            path = Path(temporary) / "dcir.xlsx"
            _write_protocol_workbook(path, plan_rows=plan, records=records)
            cell, source, metadata, _raw = self._register(path)
            reconstructed = protocol.reconstruct_protocol(
                metadata["raw"], source.nominal_capacity_mah
            )
            candidates = chargeability.detect_candidates(reconstructed)
            dcir_candidates = dcir.detect_candidates(reconstructed)
            spec = self._spec(cell)
            spec["dcir_segments"] = [
                {
                    "id": "discharge-pulse",
                    "name": "Discharge pulse",
                    "targets": [dcir_candidates[0]],
                }
            ]
            spec["computation"]["dcir"] = {
                "series": [
                    {
                        "id": "discharge-pulse-series",
                        "cell_id": cell.id,
                        "segment_id": "discharge-pulse",
                    }
                ]
            }
            result = analysis_engine.compute_dcir(self.db, spec, None)

        self.assertEqual(candidates, [])
        self.assertEqual(len(dcir_candidates), 1)
        series = result["cell_series"][0]
        self.assertEqual(series["n_measurements"], 3)
        self.assertEqual(series["direction"], "discharge")
        np.testing.assert_allclose(series["quantities"]["dcir_mohm"], [2000.0] * 3)
        self.assertEqual(result["sources"][0]["file_hashes"], [source.hash])

    def test_registered_excel_rate_fixture_is_detected_and_extracted(self):
        rates = (0.2, 0.5, 1.0)
        plan: list[list[object]] = []
        for index, rate in enumerate(rates):
            charge_step = index * 2 + 1
            discharge_step = charge_step + 1
            plan.extend(
                [
                    _plan_row(
                        charge_step,
                        "CC Chg",
                        step_time=1,
                        voltage=4.2,
                        rate=rate,
                        current=rate * 50.0,
                    ),
                    _plan_row(
                        discharge_step,
                        "CC DChg",
                        step_time=1,
                        rate=1.0,
                        current=-50.0,
                        cutoff_voltage=2.5,
                    ),
                ]
            )
        plan.append(_plan_row(7, "End"))

        records: list[dict[str, object]] = []
        base = datetime(2026, 1, 1, 12, 0, 0)
        total_min = 0.0
        data_point = 1
        for cycle, rate in enumerate(rates, start=1):
            for step_index, status, current, capacity, voltages in (
                (cycle * 2 - 1, "CC Chg", rate * 50.0, 10.0 + rate, (3.5, 4.2)),
                (cycle * 2, "CC DChg", -50.0, 8.0 + rate, (3.5, 2.5)),
            ):
                records.append(
                    _protocol_record(
                        data_point,
                        cycle,
                        step_index,
                        status,
                        0.0,
                        total_min,
                        current,
                        voltages[0],
                        0.0,
                        base + timedelta(minutes=total_min),
                    )
                )
                data_point += 1
                records.append(
                    _protocol_record(
                        data_point,
                        cycle,
                        step_index,
                        status,
                        1.0,
                        total_min + 1.0,
                        current,
                        voltages[1],
                        capacity,
                        base + timedelta(minutes=total_min + 1.0),
                    )
                )
                data_point += 1
                total_min += 1.0

        with TemporaryDirectory() as temporary:
            path = Path(temporary) / "rate-capability.xlsx"
            _write_protocol_workbook(path, plan_rows=plan, records=records)
            cell, source, metadata, _raw = self._register(path)
            reconstructed = protocol.reconstruct_protocol(
                metadata["raw"], source.nominal_capacity_mah
            )
            pairs = rate_capability.build_rate_pairs(reconstructed)
            spec = self._spec(cell)
            result = rate_capability.compute(self.db, spec, None)

        self.assertEqual(len(pairs), 3)
        self.assertEqual(
            [round(pair["charge_rate_c"], 3) for pair in pairs], [0.2, 0.5, 1.0]
        )
        families = result["cells"][0]["families"]
        self.assertEqual(families["charge"]["status"], "matched")
        self.assertEqual(families["charge"]["point_count"], 3)
        charge_points = [
            point for point in result["points"] if point["family"] == "charge"
        ]
        self.assertEqual(len(charge_points), 3)
        capacities_by_rate = {
            round(float(point["rate_c"]), 3): float(point["capacity_mah"])
            for point in charge_points
        }
        for rate, expected_capacity in (
            (0.2, 10.2),
            (0.5, 10.5),
            (1.0, 11.0),
        ):
            self.assertAlmostEqual(
                capacities_by_rate[rate], expected_capacity, places=6
            )
        self.assertEqual(families["discharge"]["status"], "not_detected")
        self.assertEqual(result["sources"][0]["file_hashes"], [source.hash])

    def test_registered_excel_without_conditions_reports_chargeability_no_match(self):
        with TemporaryDirectory() as temporary:
            path = Path(temporary) / "no-conditions.xlsx"
            _write_metadata_workbook(path)
            cell, source, metadata, _raw = self._register(path)
            spec = self._spec(cell)
            result = chargeability.compute(self.db, spec, None)

        self.assertTrue(any("protocol condition expressions" in warning for warning in metadata["protocol_warnings"]))
        self.assertEqual(result["cells"][0]["status"], "no_candidates")
        self.assertEqual(result["cells"][0]["candidate_count"], 0)
        self.assertEqual(result["cells"][0]["match_count"], 0)
        self.assertIn(
            "chargeability_no_candidates",
            {badge["kind"] for badge in result["badges"]},
        )
        self.assertEqual(result["sources"][0]["file_hashes"], [source.hash])


if __name__ == "__main__":
    unittest.main()
